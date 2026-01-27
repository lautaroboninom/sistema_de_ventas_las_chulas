import csv
import re
import sys
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from django.core.management.base import BaseCommand, CommandError
from django.db import connection, transaction

from service.models import User


ALLOWED_ROLES = {"jefe", "jefe_veedor", "admin", "recepcion"}
DEFAULT_BATCH_SIZE = 200
DEFAULT_COMMENT = "Marcado como baja desde la hoja de servicio"


def _normalize_role(value: str) -> str:
    return (value or "").strip().lower().replace(" ", "_").replace("-", "_")


def _normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (value or "").strip().lower())


def _parse_int(value) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    s = str(value).strip()
    if not s:
        return None
    try:
        return int(s)
    except Exception:
        return None


def _pick_column(headers: Sequence[str], preferred: Optional[str]) -> str:
    if not headers:
        raise CommandError("El archivo no tiene encabezados.")
    norm_map: Dict[str, str] = {_normalize_header(h): h for h in headers}
    if preferred:
        key = _normalize_header(preferred)
        if key in norm_map:
            return norm_map[key]
        raise CommandError(f"Columna '{preferred}' no encontrada en el archivo.")
    for cand in ("access_id", "os", "ingreso_id", "id"):
        key = _normalize_header(cand)
        if key in norm_map:
            return norm_map[key]
    raise CommandError("No se pudo inferir la columna de IDs (ej: access_id u OS).")


def _read_ids_from_csv(path: Path, column: Optional[str]) -> List[int]:
    ids: List[int] = []
    with path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            raise CommandError("CSV sin encabezados.")
        col = _pick_column(reader.fieldnames, column)
        for row in reader:
            val = _parse_int(row.get(col))
            if val is not None:
                ids.append(val)
    return ids


def _read_ids_from_xlsx(path: Path, column: Optional[str]) -> List[int]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise CommandError(f"No se pudo importar openpyxl: {exc}")
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            raise CommandError("XLSX sin encabezados.")
        header_strs = [str(h or "") for h in headers]
        col_name = _pick_column(header_strs, column)
        col_idx = header_strs.index(col_name)
        ids: List[int] = []
        for row in rows:
            if not row or col_idx >= len(row):
                continue
            val = _parse_int(row[col_idx])
            if val is not None:
                ids.append(val)
        return ids
    finally:
        wb.close()


def _read_ids_from_text(text: str) -> List[int]:
    tokens = re.split(r"[,\s;]+", text or "")
    ids: List[int] = []
    for tok in tokens:
        val = _parse_int(tok)
        if val is not None:
            ids.append(val)
    return ids


def _chunked(items: Sequence[int], size: int) -> Iterable[List[int]]:
    for i in range(0, len(items), size):
        yield list(items[i : i + size])


def _dash_location_id() -> Optional[int]:
    with connection.cursor() as cur:
        cur.execute("SELECT id FROM locations WHERE nombre='-' LIMIT 1")
        row = cur.fetchone()
        if row:
            return int(row[0])
        cur.execute(
            """
            INSERT INTO locations (nombre)
            SELECT %s
            WHERE NOT EXISTS (
                SELECT 1 FROM locations WHERE nombre=%s
            )
            """,
            ["-", "-"],
        )
        cur.execute("SELECT id FROM locations WHERE nombre='-' LIMIT 1")
        row = cur.fetchone()
        return int(row[0]) if row else None


def _set_audit_user(user_id: int, role: str) -> None:
    with connection.cursor() as cur:
        cur.execute("SET app.user_id = %s;", [str(user_id)])
        cur.execute("SET app.user_role = %s;", [role or ""])


def _fetch_states(ids: Sequence[int], batch_size: int) -> Dict[int, str]:
    states: Dict[int, str] = {}
    for chunk in _chunked(list(ids), batch_size):
        with connection.cursor() as cur:
            cur.execute(
                "SELECT id, estado::text FROM ingresos WHERE id = ANY(%s)",
                [chunk],
            )
            for row in cur.fetchall():
                states[int(row[0])] = (row[1] or "").lower()
    return states


class Command(BaseCommand):
    help = "Marca ingresos como baja en lote sin enviar emails."

    def add_arguments(self, parser):
        parser.add_argument("--actor-email", required=True, help="Email del usuario que registra la baja.")
        parser.add_argument("--actor-role", default="", help="Rol a registrar (default: rol del usuario).")
        parser.add_argument("--ids", nargs="*", type=int, help="Lista de OS/ids a procesar.")
        parser.add_argument("--file", help="Archivo .csv/.xlsx con ids (columna access_id/OS).")
        parser.add_argument("--column", default="", help="Columna con ids en el archivo.")
        parser.add_argument("--stdin", action="store_true", help="Leer ids desde stdin.")
        parser.add_argument("--dry-run", action="store_true", help="Solo reporta, no cambia datos.")
        parser.add_argument("--batch-size", type=int, default=DEFAULT_BATCH_SIZE, help="Tamano de lote.")

    def handle(self, *args, **opts):
        ids: List[int] = []
        if opts.get("ids"):
            ids.extend(opts["ids"])

        file_path = opts.get("file")
        if file_path:
            path = Path(file_path)
            if not path.exists():
                raise CommandError(f"No existe el archivo: {path}")
            col = opts.get("column") or None
            if path.suffix.lower() == ".csv":
                ids.extend(_read_ids_from_csv(path, col))
            elif path.suffix.lower() in (".xlsx", ".xlsm"):
                ids.extend(_read_ids_from_xlsx(path, col))
            else:
                raise CommandError("Formato no soportado. Use .csv o .xlsx.")

        if opts.get("stdin"):
            ids.extend(_read_ids_from_text(sys.stdin.read()))

        ids = sorted(set(i for i in ids if isinstance(i, int)))
        if not ids:
            raise CommandError("No se encontraron ids para procesar.")

        actor_email = (opts.get("actor_email") or "").strip().lower()
        actor_role_override = (opts.get("actor_role") or "").strip()
        user = User.objects.filter(email__iexact=actor_email, activo=True).first()
        if not user:
            raise CommandError(f"Usuario no encontrado o inactivo: {actor_email}")
        role_raw = actor_role_override or (user.rol or "")
        role_norm = _normalize_role(role_raw)
        if role_norm not in ALLOWED_ROLES:
            raise CommandError(f"Rol no permitido: {role_raw}")

        _set_audit_user(user.id, role_norm)
        dash_id = _dash_location_id()

        batch_size = int(opts.get("batch_size") or DEFAULT_BATCH_SIZE)
        states = _fetch_states(ids, batch_size)
        missing = [i for i in ids if i not in states]
        already_baja = [i for i, s in states.items() if s == "baja"]
        to_update = [i for i, s in states.items() if s and s != "baja"]

        self.stdout.write(f"Total ids: {len(ids)}")
        self.stdout.write(f"Encontrados: {len(states)}")
        self.stdout.write(f"Faltantes: {len(missing)}")
        self.stdout.write(f"Ya baja: {len(already_baja)}")
        self.stdout.write(f"A marcar baja: {len(to_update)}")

        if opts.get("dry_run"):
            return

        comment = DEFAULT_COMMENT
        updated = 0
        events = 0
        with transaction.atomic():
            for chunk in _chunked(to_update, batch_size):
                if not chunk:
                    continue
                with connection.cursor() as cur:
                    if dash_id:
                        cur.execute(
                            """
                            UPDATE ingresos
                               SET estado='baja',
                                   ubicacion_id = COALESCE(%s, ubicacion_id)
                             WHERE id = ANY(%s)
                               AND estado <> 'baja'
                            """,
                            [dash_id, chunk],
                        )
                    else:
                        cur.execute(
                            """
                            UPDATE ingresos
                               SET estado='baja'
                             WHERE id = ANY(%s)
                               AND estado <> 'baja'
                            """,
                            [chunk],
                        )
                    updated += cur.rowcount
                    cur.executemany(
                        """
                        INSERT INTO ingreso_events (ticket_id, a_estado, comentario)
                        VALUES (%s, 'baja', %s)
                        """,
                        [(cid, comment) for cid in chunk],
                    )
                    events += len(chunk)

        self.stdout.write(f"Actualizados: {updated}")
        self.stdout.write(f"Eventos creados: {events}")
