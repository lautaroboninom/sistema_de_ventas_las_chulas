from django.core.management.base import BaseCommand
from django.db import connection, transaction
from typing import Dict, List, Tuple, Any, Optional


def _norm_code(code: Optional[str]) -> Optional[str]:
    if not code:
        return None
    s = (str(code) or "").strip().upper()
    import re
    m = re.match(r"^(MG|NM|NV|CE)[^0-9]*(\d{1,6})$", s)
    if not m:
        return None
    pref, num = m.group(1), m.group(2)
    return f"{pref} {num[-4:].zfill(4)}"


def _ns_key(s: Optional[str]) -> str:
    s2 = (str(s) or "").strip().upper()
    return s2.replace(" ", "").replace("-", "")


class Command(BaseCommand):
    help = (
        "Completa devices.numero_serie a partir del MADRE, buscando devices.numero_interno en Col H (codigo) y tomando NS de Col G. "
        "Evita violar la unicidad del NS normalizado y genera un Excel de reporte."
    )

    def add_arguments(self, parser):
        parser.add_argument("--file", default="docs/Copia Para Consultar MADRE 2025.xlsx", help="Ruta del MADRE (Col G=NS, H=codigo)")
        parser.add_argument("--sheet", default="EQUILUX", help="Nombre de hoja (default: EQUILUX; si no existe usa primera)")
        parser.add_argument("--apply", action="store_true", help="Aplica cambios (default: dry-run)")
        parser.add_argument("--overwrite", action="store_true", help="Sobrescribe numero_serie aunque ya tenga valor")
        parser.add_argument("--steal", action="store_true", help="Si existe otro device con el mismo NS, limpia su numero_serie y asigna al actual")
        parser.add_argument("--bajas", default="docs/Copia de HISTORICO DE BAJAS 2024.xlsx", help="Ruta del Excel de Bajas 2024 (Col F=NS, G=codigo)")
        parser.add_argument("--out", default="docs/fill_ns_from_madre.xlsx", help="Excel de salida con resultados")

    def _load_map(self, path: str, sheet_name: Optional[str]) -> Dict[str, str]:
        out: Dict[str, str] = {}
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb[wb.sheetnames[0]]
            for row in ws.iter_rows(min_row=1):
                try:
                    ns_val = row[6].value if len(row) > 6 else None  # Col G (0-based idx 6)
                    code_val = row[7].value if len(row) > 7 else None  # Col H (0-based idx 7)
                except Exception:
                    ns_val, code_val = None, None
                code_key = _norm_code(code_val)
                ns_str = (str(ns_val).strip() if ns_val is not None else "")
                if code_key and ns_str:
                    out[code_key] = ns_str
        except Exception:
            return {}
        return out

    def _load_bajas_map(self, path: str) -> Dict[str, str]:
        out: Dict[str, str] = {}
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            for row in ws.iter_rows(min_row=1):
                try:
                    ns_val = row[5].value if len(row) > 5 else None  # Col F
                    code_val = row[6].value if len(row) > 6 else None  # Col G
                except Exception:
                    ns_val, code_val = None, None
                code_key = _norm_code(code_val)
                ns_str = (str(ns_val).strip() if ns_val is not None else "")
                if code_key and ns_str:
                    out[code_key] = ns_str
        except Exception:
            return {}
        return out

    def handle(self, *args, **opts):
        path = opts.get("file")
        sheet = opts.get("sheet")
        apply = bool(opts.get("apply"))
        out_path = opts.get("out")
        code_to_ns = self._load_map(path, sheet)
        bajas_map = self._load_bajas_map(opts.get("bajas"))

        if not code_to_ns:
            self.stdout.write("No se pudo cargar mapa codigo->NS desde el MADRE.")
            return

        # Buscar devices con numero_interno (si --overwrite, incluir todos; si no, solo vacÃ­os)
        with connection.cursor() as cur:
            if opts.get("overwrite"):
                cur.execute(
                    """
                    SELECT id, numero_interno, COALESCE(numero_serie,'')
                      FROM devices
                     WHERE NULLIF(TRIM(numero_interno),'') IS NOT NULL
                    """
                )
            else:
                cur.execute(
                    """
                    SELECT id, numero_interno, COALESCE(numero_serie,'')
                      FROM devices
                     WHERE NULLIF(TRIM(numero_interno),'') IS NOT NULL
                       AND (numero_serie IS NULL OR TRIM(numero_serie)='')
                    """
                )
            rows = cur.fetchall() or []

        to_update: List[Tuple[int, Optional[str], str]] = []  # (device_id, ns_or_None, source)
        not_found: List[Tuple[int, str, str]] = []  # (device_id, code_key, source)
        conflicts: List[Tuple[int, str, int, str]] = []  # (device_id, ns, other_id, source)

        steal = bool(opts.get("steal"))
        for rid, code, ns_current in rows:
            code_key = _norm_code(code)
            if not code_key:
                continue
            ns_new = code_to_ns.get(code_key)
            if not ns_new:
                not_found.append((int(rid), code_key, "MADRE"))
                continue
            # Si no overwrite y ya tiene valor igual normalizado, saltar
            if not opts.get("overwrite") and _ns_key(ns_current) == _ns_key(ns_new):
                continue
            # Validar conflicto de NS normalizado
            with connection.cursor() as cur:
                cur.execute(
                    """
                    SELECT id FROM devices
                     WHERE id<>%s AND REPLACE(REPLACE(UPPER(numero_serie),' ','') ,'-','')=%s
                     LIMIT 1
                    """,
                    [int(rid), _ns_key(ns_new)],
                )
                rowc = cur.fetchone()
                if rowc:
                    other_id = int(rowc[0])
                    if steal:
                        conflicts.append((int(rid), ns_new, other_id, "MADRE"))
                        to_update.append((int(other_id), None, "MADRE_CLEAR"))
                    else:
                        conflicts.append((int(rid), ns_new, other_id, "MADRE"))
                        continue
            to_update.append((int(rid), ns_new, "MADRE"))

        # Fallback: intentar completar desde BAJAS para los not_found
        if bajas_map and not_found:
            for rid, code_key, _src in list(not_found):
                ns_new = bajas_map.get(code_key)
                if not ns_new:
                    continue
                # Obtener ns_current para comparaciÃ³n si no overwrite
                ns_current = ""
                if not opts.get("overwrite"):
                    with connection.cursor() as cur:
                        rrow = None
                        try:
                            cur.execute("SELECT COALESCE(numero_serie,'') FROM devices WHERE id=%s", [rid])
                            rrow = cur.fetchone()
                        except Exception:
                            rrow = None
                    ns_current = (rrow and rrow[0]) or ""
                if not opts.get("overwrite") and _ns_key(ns_current) == _ns_key(ns_new):
                    # ya coincide, no agregar
                    continue
                with connection.cursor() as cur:
                    cur.execute(
                        """
                        SELECT id FROM devices
                         WHERE id<>%s AND REPLACE(REPLACE(UPPER(numero_serie),' ','') ,'-','')=%s
                         LIMIT 1
                        """,
                        [int(rid), _ns_key(ns_new)],
                    )
                    rowc = cur.fetchone()
                    if rowc:
                        other_id = int(rowc[0])
                        if steal:
                            conflicts.append((int(rid), ns_new, other_id, "BAJAS"))
                            to_update.append((int(other_id), None, "BAJAS_CLEAR"))
                        else:
                            conflicts.append((int(rid), ns_new, other_id, "BAJAS"))
                            continue
                to_update.append((int(rid), ns_new, "BAJAS"))
            # Filtrar not_found a los que siguen sin valor tras BAJAS
            still_not_found = []
            for rid, code_key, src in not_found:
                if code_key not in bajas_map:
                    still_not_found.append((rid, code_key, src))
            not_found = still_not_found

        # Aplicar
        if apply and to_update:
            # Validar duplicados entre los updates propuestos (mismo NS para varios devices)
            ns_groups: Dict[str, List[int]] = {}
            for rid, ns, src in to_update:
                if ns is None:
                    continue
                ns_groups.setdefault(_ns_key(ns), []).append(rid)
            # Mover duplicados a conflictos
            to_update_final: List[Tuple[int, Optional[str], str]] = []
            for rid, ns, src in to_update:
                if ns is None:
                    to_update_final.append((rid, ns, src))
                    continue
                key = _ns_key(ns)
                if len(ns_groups.get(key, [])) > 1:
                    # Marcar conflicto masivo con other_id=0
                    conflicts.append((rid, ns, 0, src))
                else:
                    to_update_final.append((rid, ns, src))

            with transaction.atomic():
                with connection.cursor() as cur:
                    # Primero limpiar los marcados con None
                    clears = [(rid, ns) for rid, ns, src in to_update_final if ns is None]
                    updates = [(rid, ns) for rid, ns, src in to_update_final if ns is not None]
                    if clears:
                        cur.executemany("UPDATE devices SET numero_serie=NULL WHERE id=%s", [(rid,) for rid, _ in clears])
                    if updates:
                        cur.executemany(
                            "UPDATE devices SET numero_serie=%s WHERE id=%s",
                            [(ns, rid) for rid, ns in updates],
                        )

        # Reporte a Excel
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "updated"
            ws1.append(["device_id", "numero_serie", "source"])
            for rid, ns, src in to_update:
                ws1.append([rid, ns if ns is not None else "<CLEARED>", src])
            ws2 = wb.create_sheet("not_found")
            ws2.append(["device_id", "code", "source"])
            for rid, code, src in not_found:
                ws2.append([rid, code, src])
            ws3 = wb.create_sheet("conflicts")
            ws3.append(["device_id", "numero_serie", "conflict_with_device_id", "source"])
            for rid, ns, other, src in conflicts:
                ws3.append([rid, ns, other, src])
            # Save
            try:
                import os
                os.makedirs(os.path.dirname(out_path), exist_ok=True)
            except Exception:
                pass
            wb.save(out_path)
            self.stdout.write(
                f"OK {'(APLICADO)' if apply else '(dry-run)'}: updates={len(to_update)} not_found={len(not_found)} conflicts={len(conflicts)} | Excel: {out_path}"
            )
        except Exception as e:
            self.stdout.write(f"No se pudo generar Excel de salida: {e}")
