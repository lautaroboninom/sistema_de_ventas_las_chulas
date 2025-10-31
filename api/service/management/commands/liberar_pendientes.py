from django.core.management.base import BaseCommand
from django.db import connection, transaction
from typing import Optional, Any, Dict, List, Tuple


def _norm_code(code: Optional[str]) -> Optional[str]:
    if not code:
        return None
    s = (str(code) or "").strip().upper()
    import re
    m = re.match(r"^(MG|NM|NV|CE)[^0-9]*(\d{1,6})$", s)
    if not m:
        return None
    pref, num = m.group(1), m.group(2)
    return f"{pref} {num[-4:].zfill(4)}"


def _norm_text(s: Optional[str]) -> str:
    try:
        import unicodedata
        s2 = (s or "").strip().lower()
        s2 = "".join(c for c in unicodedata.normalize("NFKD", s2) if not unicodedata.combining(c))
        return " ".join(s2.split())
    except Exception:
        return (s or "").strip().lower()


class Command(BaseCommand):
    help = (
        "Libera ingresos del listado 'Pendientes General' con reglas: "
        "- Corte 2022: aplica a todos (BAJAS->Desguace | MG->alquilado | resto->entregado). "
        "- Extensión 2023: aplica SOLO a MG->alquilado y a N/S prefijo UB0M010->Desguace. "
        "- BAJAS Col B='BAJA' (Excel) tiene prioridad: envía a 'Desguace'. "
        "Genera Excel con el detalle y soporta dry-run."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--bajas",
            default="docs/Copia de HISTORICO DE BAJAS 2024.xlsx",
            help="Ruta del Excel de Bajas (Col B=estado, Col G=codigo)",
        )
        parser.add_argument("--apply", action="store_true", help="Aplica cambios (default: dry-run)")
        parser.add_argument(
            "--out",
            default="docs/liberar_pendientes.xlsx",
            help="Ruta del Excel de salida con el detalle",
        )

    def _load_bajas(self, path: str) -> Dict[str, str]:
        out: Dict[str, str] = {}
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            for row in ws.iter_rows(min_row=1):
                try:
                    estado = row[1].value if len(row) > 1 else None  # Col B
                    code = row[6].value if len(row) > 6 else None   # Col G
                except Exception:
                    estado, code = None, None
                key = _norm_code(code)
                est = _norm_text(estado)
                if key and est:
                    out[key] = est
        except Exception:
            return {}
        return out

    def _ensure_location_id(self, name: str) -> Optional[int]:
        with connection.cursor() as cur:
            cur.execute("SELECT id FROM locations WHERE LOWER(nombre)=LOWER(%s) LIMIT 1", [name])
            r = cur.fetchone()
            if r:
                return int(r[0])
            try:
                cur.execute("INSERT INTO locations(nombre) VALUES (%s) RETURNING id", [name])
                return int(cur.fetchone()[0])
            except Exception:
                return None

    def handle(self, *args, **opts):
        apply = bool(opts.get("apply"))
        bajas_map = self._load_bajas(opts.get("bajas"))
        loc_desguace_id = self._ensure_location_id("Desguace")

        # Obtener ingresos en Pendientes General (loc='taller' y estado no en liberado/entregado/alquilado)
        with connection.cursor() as cur:
            cur.execute(
                """
                SELECT t.id AS ingreso_id,
                       t.estado,
                       t.fecha_ingreso,
                       COALESCE(t.fecha_creacion, t.fecha_ingreso) AS fecha_creacion,
                       t.ubicacion_id,
                       d.id AS device_id,
                       COALESCE(d.numero_interno,'') AS numero_interno,
                       COALESCE(d.numero_serie,'') AS numero_serie,
                       COALESCE(b.nombre,'') AS marca,
                       COALESCE(m.nombre,'') AS modelo
                  FROM ingresos t
                  JOIN devices d ON d.id = t.device_id
                  LEFT JOIN locations loc ON loc.id = t.ubicacion_id
                  LEFT JOIN marcas b ON b.id = d.marca_id
                  LEFT JOIN models m ON m.id = d.model_id
                 WHERE LOWER(loc.nombre) = LOWER('taller')
                   AND t.estado NOT IN ('liberado','entregado','alquilado')
                   AND DATE(COALESCE(t.fecha_ingreso, t.fecha_creacion)) <= DATE('2023-12-31')
                """
            )
            pend = cur.fetchall() or []

        updates: List[Tuple[int, str, Optional[int], str]] = []
        # (ingreso_id, new_estado or '', new_ubic_id or None, reason)

        import datetime as _dt
        cutoff_2022 = _dt.date(2022, 12, 31)
        cutoff_2023 = _dt.date(2023, 12, 31)

        for (ingreso_id, estado, f_ing, f_crea, ubic_id, device_id, numint, ns, marca, modelo) in pend:
            code = _norm_code(numint)
            reason = ""
            new_estado = None
            new_ubic = None
            # fecha de referencia (date)
            try:
                dt_ref = (f_ing or f_crea)
                if hasattr(dt_ref, 'date'):
                    dt_ref = dt_ref.date()
            except Exception:
                dt_ref = None

            ns_key = (ns or "").strip().upper()
            # 1) BAJAS = 'baja' -> Desguace (ubicacion)
            if code and bajas_map.get(code) == "baja":
                if loc_desguace_id:
                    new_ubic = loc_desguace_id
                    reason = "BAJAS->Desguace"
            else:
                # Regla UB0M010 -> Desguace (independiente del corte, si está en el lote seleccionado)
                if ns_key.startswith("UB0M010") and loc_desguace_id:
                    new_ubic = loc_desguace_id
                    reason = "NS UB0M010->Desguace"
                else:
                    # Corte 2022: todo el resto
                    if dt_ref and dt_ref <= cutoff_2022:
                        if code and code.startswith("MG "):
                            new_estado = "alquilado"
                            reason = "MG->alquilado"
                        else:
                            new_estado = "entregado"
                            reason = "default->entregado"
                    # Extensión 2023: sólo MG->alquilado
                    elif dt_ref and dt_ref <= cutoff_2023:
                        if code and code.startswith("MG "):
                            new_estado = "alquilado"
                            reason = "MG(2023)->alquilado"
                        else:
                            # fuera de alcance
                            continue
                    else:
                        # fuera de alcance
                        continue

            updates.append((int(ingreso_id), new_estado or "", new_ubic, reason))

        applied = 0
        with transaction.atomic():
            with connection.cursor() as cur:
                for ingreso_id, new_estado, new_ubic, reason in updates:
                    if not apply:
                        continue
                    if new_ubic:
                        cur.execute(
                            "UPDATE ingresos SET ubicacion_id=%s WHERE id=%s",
                            [new_ubic, ingreso_id],
                        )
                    if new_estado:
                        if new_estado == "entregado":
                            cur.execute(
                                "UPDATE ingresos SET estado='entregado', fecha_entrega=COALESCE(fecha_entrega, CURRENT_DATE) WHERE id=%s",
                                [ingreso_id],
                            )
                        elif new_estado == "alquilado":
                            cur.execute(
                                "UPDATE ingresos SET estado='alquilado', alquilado=true WHERE id=%s",
                                [ingreso_id],
                            )
                    applied += 1

        # Reporte a Excel
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "liberados"
            ws.append(["ingreso_id", "new_estado", "new_ubic", "reason"])
            for ingreso_id, new_estado, new_ubic, reason in updates:
                ws.append([ingreso_id, new_estado, new_ubic or "", reason])
            import os
            os.makedirs("docs", exist_ok=True)
            out_path = opts.get("out")
            wb.save(out_path)
            self.stdout.write(
                f"OK {'(APLICADO)' if apply else '(dry-run)'}: candidates={len(updates)} | Excel: {out_path}"
            )
        except Exception as e:
            self.stdout.write(f"No se pudo generar Excel: {e}")
