from django.core.management.base import BaseCommand
from django.db import connection, transaction
from typing import Dict, List, Tuple, Any, Optional


def _norm_key(code: Optional[str]) -> Optional[str]:
    if not code:
        return None
    s = (code or "").strip().upper()
    # Keep last 4 digits, zero-padded
    import re
    m = re.match(r"^(MG|NM|NV|CE)[^0-9]*(\d{1,6})$", s)
    if not m:
        return None
    pref, num = m.group(1), m.group(2)
    num4 = num[-4:].zfill(4)
    return f"{pref} {num4}"


def _norm_ns(ns: Optional[str]) -> str:
    s = (ns or "").strip().upper()
    return s.replace(" ", "").replace("-", "")


def _norm_text(s: Optional[str]) -> str:
    try:
        import unicodedata
        s2 = (s or "").strip().lower()
        s2 = "".join(c for c in unicodedata.normalize("NFKD", s2) if not unicodedata.combining(c))
        return " ".join(s2.split())
    except Exception:
        return (s or "").strip().lower()


def _mm_key(s: Optional[str]) -> str:
    # Marca/modelo key: remove spaces, uppercase, strip diacritics
    try:
        import unicodedata
        s2 = (s or "").strip()
        s2 = "".join(c for c in unicodedata.normalize("NFKD", s2) if not unicodedata.combining(c))
        s2 = s2.replace(" ", "").upper()
        return s2
    except Exception:
        return (s or "").strip().replace(" ", "").upper()


class Command(BaseCommand):
    help = (
        "Detecta conflictos por numero_interno normalizado (MG/NM/NV/CE ####), "+
        "propone merges cuando hay mismo numero_serie o misma marca+modelo, y "+
        "reporta casos donde difieren marca/modelo y requiere consulta a MADRE."
    )

    def add_arguments(self, parser):
        parser.add_argument("--apply", action="store_true", help="Aplica merges seguros (mismo NS o misma marca+modelo)")
        parser.add_argument("--excel", default="docs/Copia Para Consultar MADRE 2025.xlsx", help="Ruta del Excel MADRE (col H=code, E=marca, F=modelo)")
        parser.add_argument("--bajas", default="docs/Copia de HISTORICO DE BAJAS 2024.xlsx", help="Ruta del Excel de Bajas 2024 (col G=código)")
        parser.add_argument("--out", default="docs/numint_conflicts.xlsx", help="Ruta del Excel de salida con el detalle de conflictos")

    def _load_excel_map(self, path: str) -> Dict[str, Tuple[str, str]]:
        out: Dict[str, Tuple[str, str]] = {}
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            for row in ws.iter_rows(min_row=2):  # asumir encabezados en fila 1
                try:
                    marca = row[4].value if len(row) > 4 else None  # Col E
                    modelo = row[5].value if len(row) > 5 else None  # Col F
                    codigo = row[7].value if len(row) > 7 else None  # Col H
                except Exception:
                    continue
                key = _norm_key(str(codigo) if codigo is not None else "")
                if not key:
                    continue
                if key not in out:
                    out[key] = (_norm_text(str(marca) if marca is not None else ""), _norm_text(str(modelo) if modelo is not None else ""))
        except Exception:
            return {}
        return out

    def _load_bajas_codes(self, path: str) -> set[str]:
        out: set[str] = set()
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            for row in ws.iter_rows(min_row=1):
                try:
                    code = row[6].value if len(row) > 6 else None  # Col G (0-based idx 6)
                except Exception:
                    code = None
                key = _norm_key(str(code) if code is not None else "")
                if key:
                    out.add(key)
        except Exception:
            return set()
        return out

    def handle(self, *args, **opts):
        apply = bool(opts.get("apply"))
        excel_path = opts.get("excel")
        excel_map = self._load_excel_map(excel_path)
        bajas_path = opts.get("bajas")
        bajas_set = self._load_bajas_codes(bajas_path)
        out_path = opts.get("out")

        # Canonical catalog maps (marca->id, (marca,modelo)->model_id)
        brand_map: Dict[str, int] = {}
        model_map: Dict[Tuple[int, str], int] = {}
        try:
            with connection.cursor() as cur:
                cur.execute("SELECT id, nombre FROM marcas")
                for bid, bname in (cur.fetchall() or []):
                    brand_map[_mm_key(bname)] = int(bid)
                cur.execute("SELECT id, marca_id, nombre FROM models")
                for mid, bid, mname in (cur.fetchall() or []):
                    model_map[(int(bid), _mm_key(mname))] = int(mid)
        except Exception:
            brand_map = {}
            model_map = {}

        # 1) Hallar claves con conflicto por numero_interno normalizado
        with connection.cursor() as cur:
            cur.execute(
                """
                SELECT key, COUNT(*)
                  FROM (
                    SELECT UPPER(REGEXP_REPLACE(numero_interno,
                           '^(MG|NM|NV|CE)\s*(\d{1,6})$', '\\1 ' || LPAD('\\2',4,'0'))) AS key
                      FROM devices
                     WHERE numero_interno ~* '^(MG|NM|NV|CE)'
                       AND NULLIF(TRIM(numero_interno),'') IS NOT NULL
                  ) t
                 GROUP BY key
                HAVING COUNT(*) > 1
                 ORDER BY key
                """
            )
            conflict_keys = [r[0] for r in (cur.fetchall() or [])]

        if not conflict_keys:
            self.stdout.write("Sin conflictos por numero_interno.")
            return

        merges_planned: List[Tuple[int, int, str]] = []  # (from_device_id, to_device_id, reason)
        canonize_planned: List[Tuple[int, Optional[int], Optional[int], str]] = []  # (device_id, canon_brand_id, canon_model_id, why)
        unresolved_excel: List[Dict[str, Any]] = []

        for key in conflict_keys:
            # 2) Traer devices del grupo
            with connection.cursor() as cur:
                cur.execute(
                    """
                    SELECT d.id,
                           d.numero_interno,
                           d.numero_serie,
                           COALESCE(b.nombre,'') AS marca,
                           COALESCE(m.nombre,'') AS modelo,
                           (SELECT COUNT(*) FROM ingresos t WHERE t.device_id=d.id) AS cnt_ing
                      FROM devices d
                      LEFT JOIN marcas b ON b.id = d.marca_id
                      LEFT JOIN models m ON m.id = d.model_id
                     WHERE UPPER(REGEXP_REPLACE(d.numero_interno,
                           '^(MG|NM|NV|CE)\s*(\d{1,6})$', '\\1 ' || LPAD('\\2',4,'0'))) = %s
                     ORDER BY d.id
                    """,
                    [key],
                )
                devs = cur.fetchall() or []
            # Enriquecer
            items = []
            for rid, numint, ns, marca, modelo, cnt_ing in devs:
                items.append({
                    "id": int(rid),
                    "numint": numint,
                    "ns": ns or "",
                    "ns_norm": _norm_ns(ns),
                    "marca": marca or "",
                    "modelo": modelo or "",
                    "marca_k": _norm_text(marca),
                    "modelo_k": _norm_text(modelo),
                    "cnt": int(cnt_ing or 0),
                })

            # 2.a) Resolver por misma marca+modelo (normalizados: sin espacios, mayúsculas)
            groups_by_mm: Dict[Tuple[str, str], List[dict]] = {}
            for it in items:
                groups_by_mm.setdefault((_mm_key(it["marca"]), _mm_key(it["modelo"])), []).append(it)
            picked = None
            for mm, bucket in groups_by_mm.items():
                if mm != ("", "") and len(bucket) >= 2:
                    canonical = sorted(bucket, key=lambda d: (-d["cnt"], d["id"]))[0]
                    for d in bucket:
                        if d["id"] == canonical["id"]:
                            continue
                        merges_planned.append((d["id"], canonical["id"], f"same_brand_model:{mm[0]}|{mm[1]}"))
                    picked = True
                    break
            if picked:
                # Planificar canonización de marca/modelo del canónico
                try:
                    can_b_id = brand_map.get(_mm_key(canonical.get("marca")))
                    if can_b_id:
                        can_m_id = model_map.get((can_b_id, _mm_key(canonical.get("modelo"))))
                    else:
                        can_m_id = None
                    if can_b_id or can_m_id:
                        canonize_planned.append((canonical["id"], can_b_id, can_m_id, "same_brand_model"))
                except Exception:
                    pass
                continue

            # 2.b) Resolver por numero_serie igual (normalizado)
            groups_by_ns: Dict[str, List[dict]] = {}
            for it in items:
                groups_by_ns.setdefault(it["ns_norm"], []).append(it)
            merged_this_key = False
            for ns_k, bucket in groups_by_ns.items():
                if ns_k and len(bucket) >= 2:
                    canonical = sorted(bucket, key=lambda d: (-d["cnt"], d["id"]))[0]
                    for d in bucket:
                        if d["id"] == canonical["id"]:
                            continue
                        merges_planned.append((d["id"], canonical["id"], f"same_ns:{ns_k}"))
                    merged_this_key = True
            if merged_this_key:
                continue

            # 2.c) Consultar MADRE para decidir (solo reportar, no merge)
            ref = excel_map.get(key)
            if ref:
                ref_marca, ref_modelo = ref
                matches = [it for it in items if it["marca_k"] == ref_marca and it["modelo_k"] == ref_modelo]
                if not matches:
                    # ninguno coincide con MADRE -> reportar
                    unresolved_excel.append({
                        "code": key,
                        "reason": "excel_mismatch",
                        "excel_marca": ref_marca,
                        "excel_modelo": ref_modelo,
                        "found_in_bajas": key in bajas_set,
                        "devices": items,
                    })
                elif len(matches) == 1:
                    # sugerir merge hacia el que coincide con MADRE (pero no aplicar por seguridad)
                    canonical = matches[0]
                    others = [d for d in items if d["id"] != canonical["id"]]
                    for d in others:
                        unresolved_excel.append({
                            "code": key,
                            "reason": "excel_would_merge",
                            "to": canonical["id"],
                            "from": d["id"],
                            "excel_marca": ref_marca,
                            "excel_modelo": ref_modelo,
                            "found_in_bajas": key in bajas_set,
                            "devices": items,
                        })
                else:
                    # multiples coinciden con MADRE (ambiguo)
                    unresolved_excel.append({
                        "code": key,
                        "reason": "excel_ambiguous",
                        "excel_marca": ref_marca,
                        "excel_modelo": ref_modelo,
                        "found_in_bajas": key in bajas_set,
                        "devices": items,
                    })
            else:
                unresolved_excel.append({
                    "code": key,
                    "reason": "excel_not_found",
                    "found_in_bajas": key in bajas_set,
                    "devices": items,
                })

        # Resumen
        self.stdout.write(f"Merges planeados (seguros: mismo NS o misma marca+modelo): {len(merges_planned)}")
        self.stdout.write(f"Conflictos a revisar con MADRE / no resueltos: {len(unresolved_excel)}")
        self.stdout.write(f"Canonizaciones de marca/modelo planificadas: {len(canonize_planned)}")

        # Dump texto compacto
        for fr, to, why in merges_planned[:50]:
            self.stdout.write(f"  merge from={fr} -> to={to} reason={why}")
        if len(merges_planned) > 50:
            self.stdout.write(f"  ... y {len(merges_planned)-50} más")

        for row in unresolved_excel[:10]:
            code = row.get("code")
            reason = row.get("reason")
            excel_marca = row.get("excel_marca")
            excel_modelo = row.get("excel_modelo")
            in_bajas = row.get("found_in_bajas")
            devs = row.get("devices") or []
            self.stdout.write(f"[UNRESOLVED] code={code} reason={reason} excel_marca={excel_marca} excel_modelo={excel_modelo} found_in_bajas={in_bajas}")
            for it in devs:
                self.stdout.write(f"   - id={it['id']} ns='{it['ns']}' marca='{it['marca']}' modelo='{it['modelo']}' cnt={it['cnt']}")
        if len(unresolved_excel) > 10:
            self.stdout.write(f"  ... y {len(unresolved_excel)-10} más")

        # Aplicación de merges seguros (opcional)
        if apply and merges_planned:
            with transaction.atomic():
                with connection.cursor() as cur:
                    for fr, to, why in merges_planned:
                        # Reasignar ingresos
                        cur.execute("UPDATE ingresos SET device_id=%s WHERE device_id=%s", [to, fr])
                        # Limpiar numero_interno en duplicado para evitar colisiones
                        cur.execute("UPDATE devices SET numero_interno=NULL WHERE id=%s", [fr])
                    # Aplicar canonizaciones
                    for dev_id, cbid, cmid, why in canonize_planned:
                        sets = []
                        params: List[Any] = []
                        if cbid:
                            sets.append("marca_id=%s")
                            params.append(cbid)
                        if cmid:
                            sets.append("model_id=%s")
                            params.append(cmid)
                        if sets:
                            params.append(dev_id)
                            cur.execute(f"UPDATE devices SET {', '.join(sets)} WHERE id=%s", params)
            self.stdout.write("APLICADO: merges seguros ejecutados")

        # Exportar a Excel (detalle)
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "merges"
            ws1.append(["from_device_id", "to_device_id", "reason"]) 
            for fr, to, why in merges_planned:
                ws1.append([fr, to, why])

            ws2 = wb.create_sheet("unresolved")
            ws2.append([
                "code", "reason", "found_in_bajas", "excel_marca", "excel_modelo",
                "device_id", "numero_interno", "numero_serie", "marca", "modelo", "ingresos_count"
            ])
            for row in unresolved_excel:
                code = row.get("code")
                reason = row.get("reason")
                finb = row.get("found_in_bajas")
                em = row.get("excel_marca")
                eMo = row.get("excel_modelo")
                for it in (row.get("devices") or []):
                    ws2.append([
                        code,
                        reason,
                        bool(finb),
                        em,
                        eMo,
                        it.get("id"),
                        it.get("numint"),
                        it.get("ns"),
                        it.get("marca"),
                        it.get("modelo"),
                        it.get("cnt"),
                    ])
            # Guardar
            try:
                import os
                os.makedirs(os.path.dirname(out_path), exist_ok=True)
            except Exception:
                pass
            wb.save(out_path)
            self.stdout.write(f"EXCEL generado: {out_path}")
        except Exception as e:
            self.stdout.write(f"No se pudo generar Excel de salida: {e}")
