from django.core.management.base import BaseCommand
from django.db import connection, transaction
from typing import Dict, List, Tuple, Optional, Any, Set
import os
import csv
import re


class Command(BaseCommand):
    help = (
        "Sincroniza estado de alquiler a partir de ALQS.xlsx, ALQMG.xlsx y Copia Madre (hoja EQUILUX). "
        "Marca como alquilados los equipos listados y setea 'alquiler_a' en el Ãºltimo ingreso y en devices. "
        "Genera reporte de MG no alquilados (excluyendo EstanterÃ­a de Alquiler y Desguace)."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Simula cambios y genera reportes sin escribir en BD (default)",
        )
        parser.add_argument(
            "--apply",
            action="store_true",
            help="Aplica cambios en una transacciÃ³n atÃ³mica",
        )
        parser.add_argument(
            "--limit",
            type=int,
            default=None,
            help="LÃ­mite mÃ¡ximo de filas a aplicar (para pruebas)",
        )
        parser.add_argument(
            "--docs-dir",
            default=None,
            help="Directorio de salida para reportes (por defecto autodetecta 'docs' o '../docs')",
        )
        parser.add_argument(
            "--add-mg",
            action="append",
            default=[],
            help=(
                "Agregar manualmente MG a marcar como alquilados. "
                "Puede repetirse. Acepta variantes (ej: 'MG1716', 'MG 1716', 'NM 0233')."
            ),
        )
        parser.add_argument(
            "--add-code",
            action="append",
            default=[],
            help=(
                "Agregar manualmente cÃ³digos internos (MG|NM|CE) a marcar como alquilados. "
                "Puede repetirse."
            ),
        )
        parser.add_argument(
            "--rename-code",
            action="append",
            default=[],
            help=(
                "Renombra cÃ³digos internos en devices.n_de_control. Formato FROM:TO (respeta prefijo). "
                "Puede repetirse."
            ),
        )
        parser.add_argument(
            "--alquiler-a",
            dest="alquiler_a_manual",
            default=None,
            help=(
                "Texto para 'alquiler_a' al aplicar a los MG agregados manualmente. "
                "Si no se especifica, no se modifica ese campo."
            ),
        )

    # ---- Utilidades de normalizaciÃ³n ----
    @staticmethod
    def _norm_text(s: Optional[str]) -> str:
        try:
            import unicodedata
            s2 = (s or "").strip()
            s2 = "".join(
                c for c in unicodedata.normalize("NFKD", s2)
                if not unicodedata.combining(c)
            )
            # Compactar espacios
            return " ".join(s2.split())
        except Exception:
            return (s or "").strip()

    @staticmethod
    def _norm_lower_nodiac(s: Optional[str]) -> str:
        try:
            import unicodedata
            s2 = (s or "").strip().lower()
            s2 = "".join(
                c for c in unicodedata.normalize("NFKD", s2)
                if not unicodedata.combining(c)
            )
            return " ".join(s2.split())
        except Exception:
            return (s or "").strip().lower()

    @staticmethod
    def _norm_ns(ns: Optional[str]) -> str:
        # Normaliza N/S para matching por NS (quita espacios y guiones internos y upper)
        s = (ns or "").strip()
        s = s.upper()
        s = s.replace(" ", "").replace("-", "")
        return s

    @staticmethod
    def _norm_mg(mg: Optional[str]) -> Optional[str]:
        if not mg:
            return None
        s = (mg or "").upper().strip()
        # Extraer dÃ­gitos
        m = re.search(r"(\d{1,4})$", s)
        if not m:
            return None
        num = m.group(1)
        if len(num) < 4:
            num = num.zfill(4)
        out = f"MG {num}"
        # ValidaciÃ³n final
        return out if re.match(r"^MG\s\d{4}$", out) else None

    @staticmethod
    def _ensure_docs_dir(path: str) -> None:
        try:
            os.makedirs(path, exist_ok=True)
        except Exception:
            pass

    @staticmethod
    def _norm_stockcode(code: Optional[str]) -> Optional[str]:
        """Normaliza cÃ³digos internos de stock (MG|NM|CE) a formato 'XX ####'."""
        if not code:
            return None
        s = (code or "").upper().strip()
        m = re.match(r"^(MG|NM|CE)\s*(\d{1,4})$", s)
        if not m:
            m2 = re.match(r"^(MG|NM|CE)[^0-9]*(\d{1,4})$", s)
            if not m2:
                return None
            pref, num = m2.group(1), m2.group(2)
        else:
            pref, num = m.group(1), m.group(2)
        return f"{pref} {num.zfill(4)}"

    @staticmethod
    def _fmt_dt(x: Any) -> str:
        try:
            if x is None:
                return ""
            # Convertir a naive string legible
            if hasattr(x, "tzinfo"):
                try:
                    x = x.replace(tzinfo=None)
                except Exception:
                    return str(x)
            # Formato estÃ¡ndar corto
            try:
                return x.strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                return str(x)
        except Exception:
            return ""

    # ---- Lectura de Excel ----
    def _load_ns_to_code(self, madre_path: str, sheet_name: str = "EQUILUX") -> Dict[str, str]:
        from openpyxl import load_workbook
        out: Dict[str, str] = {}
        try:
            wb = load_workbook(madre_path, read_only=True, data_only=True)
        except Exception as e:
            self.stderr.write(f"No se pudo abrir '{madre_path}': {e}")
            return out
        if sheet_name not in wb.sheetnames:
            ws = wb[wb.sheetnames[0]]
        else:
            ws = wb[sheet_name]
        # Col G = 7 (NS), Col H = 8 (MG)
        for row in ws.iter_rows(min_row=1):
            try:
                ns_val = row[6].value if len(row) >= 7 else None
                mg_val = row[7].value if len(row) >= 8 else None
            except Exception:
                ns_val = None
                mg_val = None
            if not ns_val or not mg_val:
                continue
            ns_key = self._norm_ns(str(ns_val))
            mg_norm = self._norm_mg(str(mg_val))
            if ns_key and mg_norm:
                # Ãºltima gana, pero preservamos consistencia simple
                out[ns_key] = mg_norm
        return out

    def _load_alq_rows(self, xlsx_path: str) -> List[Dict[str, str]]:
        from openpyxl import load_workbook
        rows: List[Dict[str, str]] = []
        try:
            wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        except Exception as e:
            self.stderr.write(f"No se pudo abrir '{xlsx_path}': {e}")
            return rows
        ws = wb[wb.sheetnames[0]]
        # Col A = cliente (1), Col D = NS (4)
        for row in ws.iter_rows(min_row=1):
            try:
                cliente = row[0].value if len(row) >= 1 else None
                ns_val = row[3].value if len(row) >= 4 else None
            except Exception:
                cliente = None
                ns_val = None
            cliente = (str(cliente).strip() if cliente is not None else "")
            ns_raw = (str(ns_val).strip() if ns_val is not None else "")
            if not cliente and not ns_raw:
                continue
            # HeurÃ­stica simple para saltar encabezado: si 'cliente' incluye 'cliente' y ns parece literal 'ns'
            hdr = self._norm_lower_nodiac(cliente)
            if hdr.startswith("cliente"):
                continue
            rows.append({
                "cliente": cliente,
                "ns_raw": ns_raw,
                "source": os.path.basename(xlsx_path),
            })
        return rows

    # ---- Consultas BD ----
    def _fetch_device_by_mg_or_ns(self, cur, mg: Optional[str], ns_raw: str) -> Tuple[List[Dict[str, Any]], str]:
        """Devuelve lista de devices candidatos y el criterio usado."""
        # 1) Por MG exacto (en numero_interno o numero_serie)
        if mg:
            cur.execute(
                """
                SELECT id, customer_id, numero_interno, numero_serie
                  FROM devices
                 WHERE (numero_interno = %s)
                    OR (numero_serie = %s)
                """,
                [mg, mg],
            )
            rows = [
                {"id": r[0], "customer_id": r[1], "numero_interno": r[2], "numero_serie": r[3]}
                for r in (cur.fetchall() or [])
            ]
            if rows:
                return rows, "mg"
        # 2) Por NS (robusto: UPPER y sin guiones/espacios)
        ns_key = self._norm_ns(ns_raw)
        cur.execute(
            """
            SELECT id, customer_id, numero_interno, numero_serie
              FROM devices
             WHERE UPPER(numero_serie) = UPPER(%s)
                OR REPLACE(REPLACE(UPPER(numero_serie),' ',''),'-','') = %s
            """,
            [ns_raw, ns_key],
        )
        rows = [
            {"id": r[0], "customer_id": r[1], "numero_interno": r[2], "numero_serie": r[3]}
            for r in (cur.fetchall() or [])
        ]
        return rows, "ns"

    # SelecciÃ³n canÃ³nica y detecciÃ³n de conflictos reales (por MG/NS, no por device_id)
    def _pick_canonical_device(
        self,
        devices: List[Dict[str, Any]],
        mg_res: Optional[str],
        ns_key: str,
    ) -> Tuple[Optional[Dict[str, Any]], bool, Dict[str, Any]]:
        if not devices:
            return None, False, {"reason": "no_devices"}

        def dev_mg(d: Dict[str, Any]) -> Optional[str]:
            return self._norm_mg(d.get("numero_interno")) or (
                self._norm_mg(d.get("numero_serie")) if d.get("numero_serie") else None
            )

        def dev_ns(d: Dict[str, Any]) -> str:
            return self._norm_ns(d.get("numero_serie")) if d.get("numero_serie") else ""

        # Clave de equipo por contexto: si hay MG mapeado, usar MG; si no, usar NS normalizado
        if mg_res:
            keys = [dev_mg(d) or "" for d in devices]
        else:
            keys = [dev_ns(d) or (dev_mg(d) or "") for d in devices]

        unique_keys = set([k for k in keys if k])

        # Elegir canÃ³nico: preferir el que matchee mg_res (si existe), sino mayor device.id
        candidates = devices
        if mg_res:
            eq_mg = [d for d in devices if dev_mg(d) == mg_res]
            if eq_mg:
                candidates = eq_mg
        canonical = max(candidates, key=lambda d: int(d.get("id")))

        # Conflicto real si hay mÃ¡s de una clave de equipo entre candidatos
        is_conflict = len(unique_keys) > 1

        # O si hay mg_res pero ninguna clave coincide (mismatch entre mapeo y BD)
        if mg_res and (mg_res not in unique_keys):
            is_conflict = True

        meta = {
            "unique_keys": list(unique_keys),
            "mg_res": mg_res,
            "ns_key": ns_key,
            "canonical_id": int(canonical.get("id")),
            "reason": "multi_keys" if len(unique_keys) > 1 else ("mg_mismatch" if mg_res and mg_res not in unique_keys else "ok"),
        }
        return canonical, is_conflict, meta

    def _fetch_last_ingreso_id(self, cur, device_id: int) -> Optional[int]:
        cur.execute(
            """
            SELECT id
              FROM ingresos
             WHERE device_id = %s
             ORDER BY COALESCE(fecha_ingreso, fecha_creacion) DESC, id DESC
             LIMIT 1
            """,
            [device_id],
        )
        row = cur.fetchone()
        return int(row[0]) if row else None

    def _fetch_last_location(self, cur, device_id: int) -> Optional[str]:
        cur.execute(
            """
            SELECT COALESCE(l.nombre,'')
              FROM ingresos t
              LEFT JOIN locations l ON l.id = t.ubicacion_id
             WHERE t.device_id = %s
             ORDER BY COALESCE(t.fecha_ingreso, t.fecha_creacion) DESC, t.id DESC
             LIMIT 1
            """,
            [device_id],
        )
        row = cur.fetchone()
        return (row[0] or "") if row else None

    def _is_excluded_location(self, name: Optional[str]) -> bool:
        # Excluir EstanterÃ­a de Alquiler o Desguace (normalizando)
        key = self._norm_lower_nodiac(name)
        variants = {
            self._norm_lower_nodiac("EstanterÃ­a de Alquiler"),
            self._norm_lower_nodiac("Estanteria de Alquiler"),
            self._norm_lower_nodiac("EstanterÃ­a alquiler"),
            self._norm_lower_nodiac("Estanteria alquiler"),
            self._norm_lower_nodiac("Desguace"),
        }
        return key in variants

    # ---- Reportes ----
    def _write_backup_csv(self, path: str, headers: List[str], rows: List[List[Any]]) -> None:
        try:
            with open(path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(headers)
                for r in rows:
                    w.writerow(r)
        except Exception as e:
            self.stderr.write(f"No se pudo escribir backup '{path}': {e}")

    def _write_result_xlsx(self, path: str, tabs: Dict[str, List[List[Any]]]) -> None:
        try:
            from openpyxl import Workbook
            wb = Workbook()
            # Elegir primera hoja segun primer key o 'aplicados'
            if tabs:
                first_name = next(iter(tabs.keys()))
            else:
                first_name = "aplicados"
            ws0 = wb.active
            ws0.title = (first_name[:31] or "sheet")
            for row in tabs.get(first_name, []):
                ws0.append(row)
            # Resto de hojas
            for name, rows in tabs.items():
                if name == first_name:
                    continue
                ws = wb.create_sheet(title=name[:31] or "sheet")
                for r in rows:
                    ws.append(r)
            wb.save(path)
        except Exception as e:
            self.stderr.write(f"No se pudo escribir reporte '{path}': {e}")

    def _pick_docs_dir(self, user_docs_dir: Optional[str]) -> str:
        # 1) respeta parÃ¡metro explÃ­cito
        if user_docs_dir:
            return user_docs_dir
        # 2) si existe ./docs
        if os.path.isdir(os.path.join("docs")):
            return os.path.join("docs")
        # 3) si existe ../docs (ej: se ejecuta desde api/)
        if os.path.isdir(os.path.join("..", "docs")):
            return os.path.join("..", "docs")
        # 4) fallback ./docs
        return os.path.join("docs")

    def _resolve_input(self, base_docs: str, filename: str) -> str:
        p1 = os.path.join(base_docs, filename)
        if os.path.exists(p1):
            return p1
        # fallback alterno si el base_docs fue incorrecto
        alt = os.path.join("..", base_docs, filename) if not base_docs.startswith("..") else os.path.join(base_docs[3:], filename)
        return alt if os.path.exists(alt) else p1

    def _ensure_location_id(self, cur, target_name: str) -> int:
        # Buscar por igualdad case-insensitive
        cur.execute(
            "SELECT id FROM locations WHERE LOWER(nombre)=LOWER(%s) LIMIT 1",
            [target_name],
        )
        r = cur.fetchone()
        if r:
            return int(r[0])
        # Buscar variantes sin acentos/espacios mÃºltiples
        def _norm(txt: str) -> str:
            try:
                import unicodedata
                s = (txt or "").strip().lower()
                s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
                return " ".join(s.split())
            except Exception:
                return (txt or "").strip().lower()
        try:
            cur.execute("SELECT id, nombre FROM locations")
            rows = cur.fetchall() or []
            tgt = _norm(target_name)
            for lid, lname in rows:
                if _norm(lname) == tgt:
                    return int(lid)
        except Exception:
            pass
        # Crear si no existe
        cur.execute("INSERT INTO locations(nombre) VALUES (%s) RETURNING id", [target_name])
        return int(cur.fetchone()[0])

    def handle(self, *args, **opts):
        dry = True
        if opts.get("apply"):
            dry = False
        if opts.get("dry_run"):
            dry = True

        docs_dir = self._pick_docs_dir(opts.get("docs_dir"))
        self._ensure_docs_dir(docs_dir)

        # Rutas por defecto
        alqs_path = self._resolve_input(docs_dir, "ALQS.xlsx")
        alqmg_path = self._resolve_input(docs_dir, "ALQMG.xlsx")
        madre_path = self._resolve_input(docs_dir, "Copia Para Consultar MADRE 2025.xlsx")

        self.stdout.write("Cargando mapeo NS->codigo desde MADRE (EQUILUX)...")
        ns_to_code = self._load_ns_to_code(madre_path, sheet_name="EQUILUX")
        self.stdout.write(f"Mapeos NS->codigo cargados: {len(ns_to_code)}")

        # Cargar filas de ALQ*
        alq_rows = []
        alq_rows.extend(self._load_alq_rows(alqs_path))
        alq_rows.extend(self._load_alq_rows(alqmg_path))

        if not alq_rows:
            self.stderr.write("No se encontraron filas en ALQS/ALQMG. Abortando.")
            return

        # Preparar resultados
        aplicados_rows: List[List[Any]] = [[
            "device_id", "ingreso_id", "criterio", "MG", "NS", "cliente",
            "alquilado_dev_prev", "alquiler_a_dev_prev",
            "alquilado_ing_prev", "alquiler_a_ing_prev",
        ]]
        conflictos_rows: List[List[Any]] = [[
            "source", "cliente", "NS", "NS_normalizado", "criterio_intentado",
            "MG_resuelto", "coincidencias", "detalle_candidatos"
        ]]
        conflictos_detalle_rows: List[List[Any]] = [[
            # Datos de entrada
            "source", "cliente_input", "NS_raw", "NS_normalizado", "MG_resuelto", "criterio_intentado", "coincidencias_total",
            # Candidato (device)
            "device_id", "customer_id", "cliente_bd", "marca", "modelo", "tipo_equipo",
            "mg_device", "ns_device", "alquilado_device", "alquiler_a_device",
            # Ãšltimo ingreso
            "last_ingreso_id", "last_estado", "last_fecha_ingreso", "last_fecha_creacion",
            "last_ubicacion_id", "last_ubicacion_nombre", "alquilado_ingreso", "alquiler_a_ingreso",
        ]]
        no_encontrados_rows: List[List[Any]] = [["source", "cliente", "NS", "MG_resuelto"]]
        mg_no_alquilados_rows: List[List[Any]] = [["device_id", "MG", "NS", "ultima_ubicacion", "cliente_actual"]]

        # Backups (si vamos a aplicar)
        backup_devices: List[List[Any]] = [["device_id", "alquilado", "alquiler_a", "numero_interno", "numero_serie"]]
        backup_ingresos: List[List[Any]] = [["ingreso_id", "alquilado", "alquiler_a"]]

        # Para filtrar MG ya tratados
        mg_alquilados_aplicados: Set[str] = set()

        limit = opts.get("limit")

        with transaction.atomic():
            with connection.cursor() as cur:
                applied_count = 0
                for row in alq_rows:
                    cliente = self._norm_text(row.get("cliente"))
                    ns_raw = (row.get("ns_raw") or "").strip()
                    ns_key = self._norm_ns(ns_raw)
                    mg_res = self._norm_mg(ns_to_code.get(ns_key)) if ns_key in ns_to_code else None

                    devices, criterio = self._fetch_device_by_mg_or_ns(cur, mg_res, ns_raw)

                    if not devices:
                        no_encontrados_rows.append([
                            row.get("source"), cliente, ns_raw, mg_res or "",
                        ])
                        continue
                    # Elegir canÃ³nico por NS/MG (no por device_id) y evaluar conflicto real
                    canonical, is_conflict, meta = self._pick_canonical_device(devices, mg_res, ns_key)
                    # Enriquecer detalle de candidatos (si hay conflicto o si se quiere auditorÃ­a amplia)
                    if is_conflict:
                        cand_lines: List[str] = []
                        for d in devices:
                            did = int(d.get("id"))
                            # Datos bÃ¡sicos del device
                            cur.execute(
                                """
                                SELECT d.numero_interno, d.numero_serie,
                                       COALESCE(c.razon_social,'') AS cliente,
                                       c.id AS customer_id,
                                       COALESCE(b.nombre,'') AS marca,
                                       COALESCE(m.nombre,'') AS modelo,
                                       COALESCE(m.tipo_equipo,'') AS tipo_equipo,
                                       d.alquilado, d.alquiler_a
                                  FROM devices d
                                  JOIN customers c ON c.id = d.customer_id
                                  LEFT JOIN marcas b ON b.id = d.marca_id
                                  LEFT JOIN models m ON m.id = d.model_id
                                  WHERE d.id=%s
                                """,
                                [did],
                            )
                            dr = cur.fetchone() or [None, None, "", None, "", "", "", None, None]
                            # Ãšltimo ingreso detallado
                            cur.execute(
                                """
                                SELECT t.id, t.estado, t.fecha_ingreso, t.fecha_creacion,
                                       l.id AS ubic_id, COALESCE(l.nombre,'') AS ubic_nombre,
                                       t.alquilado, t.alquiler_a
                                  FROM ingresos t
                                  LEFT JOIN locations l ON l.id = t.ubicacion_id
                                 WHERE t.device_id = %s
                                 ORDER BY COALESCE(t.fecha_ingreso, t.fecha_creacion) DESC, t.id DESC
                                 LIMIT 1
                                """,
                                [did],
                            )
                            li = cur.fetchone() or [None, None, None, None, None, "", None, None]
                            ubic = li[5] or ""
                            cand_lines.append(
                                f"id={did}|MG={dr[0] or ''}|NS={dr[1] or ''}|Cliente={dr[2] or ''}|Ubic={ubic}"
                            )
                            conflictos_detalle_rows.append([
                                row.get("source"), cliente, ns_raw, ns_key, mg_res or "", criterio, len(devices),
                                did, dr[3], dr[2], dr[4], dr[5], dr[6],
                                dr[0], dr[1], dr[7], dr[8],
                                li[0], li[1], self._fmt_dt(li[2]), self._fmt_dt(li[3]), li[4], li[5], li[6], li[7],
                            ])
                        conflictos_rows.append([
                            row.get("source"), cliente, ns_raw, ns_key, criterio,
                            mg_res or "", len(devices), " ; ".join(cand_lines)
                        ])

                    # Usar el canÃ³nico seleccionado (por device_id mayor o por MG_res)
                    dev = canonical or devices[0]
                    device_id = int(dev["id"]) if isinstance(dev, dict) else int(dev[0])
                    n_de_control = dev.get("numero_interno") if isinstance(dev, dict) else None
                    numero_serie = dev.get("numero_serie") if isinstance(dev, dict) else None

                    # Determinar MG final (si no lo tenemos y el device ya lo tiene vÃ¡lido)
                    mg_final = mg_res or self._norm_mg(n_de_control) or (self._norm_mg(numero_serie) if numero_serie else None)

                    # Ãšltimo ingreso: segÃºn el device canÃ³nico (regla: mayor device_id representa Ãºltimo OS)
                    ingreso_id = self._fetch_last_ingreso_id(cur, device_id)

                    # Backups actuales
                    cur.execute(
                        "SELECT alquilado, alquiler_a, numero_interno, numero_serie FROM devices WHERE id=%s",
                        [device_id],
                    )
                    dprev = cur.fetchone() or [None, None, None, None]
                    backup_devices.append([device_id, dprev[0], dprev[1], dprev[2], dprev[3]])

                    alquilado_ing_prev = None
                    alquiler_a_ing_prev = None
                    if ingreso_id is not None:
                        cur.execute(
                            "SELECT alquilado, alquiler_a FROM ingresos WHERE id=%s",
                            [ingreso_id],
                        )
                        iprev = cur.fetchone() or [None, None]
                        alquilado_ing_prev, alquiler_a_ing_prev = iprev[0], iprev[1]
                        backup_ingresos.append([ingreso_id, alquilado_ing_prev, alquiler_a_ing_prev])

                    # AplicaciÃ³n o simulaciÃ³n
                    if not dry:
                        # Completar MG en device si falta y lo tenemos
                        if mg_final and (dprev[2] or "").strip() != mg_final:
                            cur.execute(
                                "UPDATE devices SET numero_interno=%s WHERE id=%s",
                                [mg_final, device_id],
                            )
                        # Marcar alquilado en device
                        cur.execute(
                            "UPDATE devices SET alquilado=true, alquiler_a=%s WHERE id=%s",
                            [cliente, device_id],
                        )
                        # Marcar alquilado solo en el Ãºltimo ingreso
                        if ingreso_id is not None:
                            cur.execute(
                                "UPDATE ingresos SET alquilado=true, alquiler_a=%s WHERE id=%s",
                                [cliente, ingreso_id],
                            )

                    aplicados_rows.append([
                        device_id, ingreso_id or "", criterio, mg_final or "", ns_raw, cliente,
                        dprev[0], dprev[1], alquilado_ing_prev, alquiler_a_ing_prev,
                    ])
                    if mg_final:
                        mg_alquilados_aplicados.add(mg_final)
                    applied_count += 1

                # Procesar codigos agregados manualmente (MG/NM/CE) sin convertir prefijos
                add_code_list = opts.get("add_code") or []
                alquiler_a_manual = opts.get("alquiler_a_manual")
                if add_code_list:
                    loc_id = self._ensure_location_id(cur, "Estanter de Alquiler")
                    for code_in in add_code_list:
                        code_norm = self._norm_stockcode(str(code_in))
                        if not code_norm:
                            conflictos_rows.append([
                                "manual", alquiler_a_manual or "", code_in, "", "code",
                                "", 0, "Cdigo interno invlido"
                            ])
                            continue
                        cur.execute(
                            """
                            SELECT id, customer_id, numero_interno, numero_serie
                              FROM devices
                             WHERE (numero_interno = %s) OR (numero_serie = %s)
                            """,
                            [code_norm, code_norm],
                        )
                        devs = [
                            {"id": r[0], "customer_id": r[1], "numero_interno": r[2], "numero_serie": r[3]}
                            for r in (cur.fetchall() or [])
                        ]
                        if not devs:
                            no_encontrados_rows.append(["manual", alquiler_a_manual or "", code_in, code_norm])
                            continue
                        canonical, is_conflict, meta = self._pick_canonical_device(devs, code_norm, "")
                        dev = canonical or devs[0]
                        device_id = int(dev.get("id"))
                        ingreso_id = self._fetch_last_ingreso_id(cur, device_id)
                        cur.execute(
                            "SELECT alquilado, alquiler_a, numero_interno, numero_serie FROM devices WHERE id=%s",
                            [device_id],
                        )
                        dprev = cur.fetchone() or [None, None, None, None]
                        backup_devices.append([device_id, dprev[0], dprev[1], dprev[2], dprev[3]])
                        alquilado_ing_prev = None
                        alquiler_a_ing_prev = None
                        if ingreso_id is not None:
                            cur.execute(
                                "SELECT alquilado, alquiler_a FROM ingresos WHERE id=%s",
                                [ingreso_id],
                            )
                            iprev = cur.fetchone() or [None, None]
                            alquilado_ing_prev, alquiler_a_ing_prev = iprev[0], iprev[1]
                            backup_ingresos.append([ingreso_id, alquilado_ing_prev, alquiler_a_ing_prev])
                        if not dry:
                            if (dprev[2] or "").strip() != code_norm:
                                cur.execute(
                                    "UPDATE devices SET numero_interno=%s WHERE id=%s",
                                    [code_norm, device_id],
                                )
                            if alquiler_a_manual is not None:
                                cur.execute(
                                    "UPDATE devices SET alquilado=true, alquiler_a=%s WHERE id=%s",
                                    [alquiler_a_manual, device_id],
                                )
                            else:
                                cur.execute(
                                    "UPDATE devices SET alquilado=true WHERE id=%s",
                                    [device_id],
                                )
                            if ingreso_id is not None:
                                if alquiler_a_manual is not None:
                                    cur.execute(
                                        "UPDATE ingresos SET alquilado=true, alquiler_a=%s, ubicacion_id=%s WHERE id=%s",
                                        [alquiler_a_manual, loc_id, ingreso_id],
                                    )
                                else:
                                    cur.execute(
                                        "UPDATE ingresos SET alquilado=true, ubicacion_id=%s WHERE id=%s",
                                        [loc_id, ingreso_id],
                                    )
                        aplicados_rows.append([
                            device_id, ingreso_id or "", "code_manual", code_norm, "", alquiler_a_manual or "",
                            dprev[0], dprev[1], alquilado_ing_prev, alquiler_a_ing_prev,
                        ])
                        applied_count += 1

                # Procesar MG agregados manualmente
                add_mg_list = opts.get("add_mg") or []
                alquiler_a_manual = opts.get("alquiler_a_manual")
                if add_mg_list:
                    # Preparar ubicaciÃ³n canÃ³nica
                    loc_id = self._ensure_location_id(cur, "EstanterÃ­a de Alquiler")
                    for mg_in in add_mg_list:
                        mg_norm = self._norm_mg(str(mg_in))
                        if not mg_norm:
                            conflictos_rows.append([
                                "manual", alquiler_a_manual or "", mg_in, "", "mg",
                                "", 0, "MG invÃ¡lido"
                            ])
                            continue
                        # Buscar devices por MG exacto en numero_interno o numero_serie
                        cur.execute(
                            """
                            SELECT id, customer_id, numero_interno, numero_serie
                              FROM devices
                             WHERE (numero_interno = %s) OR (numero_serie = %s)
                            """,
                            [mg_norm, mg_norm],
                        )
                        devs = [
                            {"id": r[0], "customer_id": r[1], "numero_interno": r[2], "numero_serie": r[3]}
                            for r in (cur.fetchall() or [])
                        ]
                        if not devs:
                            no_encontrados_rows.append(["manual", alquiler_a_manual or "", mg_in, mg_norm])
                            continue
                        canonical, is_conflict, meta = self._pick_canonical_device(devs, mg_norm, "")
                        dev = canonical or devs[0]
                        device_id = int(dev.get("id"))
                        ingreso_id = self._fetch_last_ingreso_id(cur, device_id)

                        # Backup previos
                        cur.execute(
                            "SELECT alquilado, alquiler_a, numero_interno, numero_serie FROM devices WHERE id=%s",
                            [device_id],
                        )
                        dprev = cur.fetchone() or [None, None, None, None]
                        backup_devices.append([device_id, dprev[0], dprev[1], dprev[2], dprev[3]])
                        alquilado_ing_prev = None
                        alquiler_a_ing_prev = None
                        if ingreso_id is not None:
                            cur.execute(
                                "SELECT alquilado, alquiler_a FROM ingresos WHERE id=%s",
                                [ingreso_id],
                            )
                            iprev = cur.fetchone() or [None, None]
                            alquilado_ing_prev, alquiler_a_ing_prev = iprev[0], iprev[1]
                            backup_ingresos.append([ingreso_id, alquilado_ing_prev, alquiler_a_ing_prev])

                        # Aplicar
                        if not dry:
                            # Asegurar MG en device
                            if (dprev[2] or "").strip() != mg_norm:
                                cur.execute("UPDATE devices SET numero_interno=%s WHERE id=%s", [mg_norm, device_id])
                            # Marcar alquilado en device (no forzar alquiler_a si no se pasa)
                            if alquiler_a_manual is not None:
                                cur.execute(
                                    "UPDATE devices SET alquilado=true, alquiler_a=%s WHERE id=%s",
                                    [alquiler_a_manual, device_id],
                                )
                            else:
                                cur.execute(
                                    "UPDATE devices SET alquilado=true WHERE id=%s",
                                    [device_id],
                                )
                            # Marcar Ãºltimo ingreso alquilado y setear ubicaciÃ³n a EstanterÃ­a de Alquiler
                            if ingreso_id is not None:
                                if alquiler_a_manual is not None:
                                    cur.execute(
                                        "UPDATE ingresos SET alquilado=true, alquiler_a=%s, ubicacion_id=%s WHERE id=%s",
                                        [alquiler_a_manual, loc_id, ingreso_id],
                                    )
                                else:
                                    cur.execute(
                                        "UPDATE ingresos SET alquilado=true, ubicacion_id=%s WHERE id=%s",
                                        [loc_id, ingreso_id],
                                    )

                        aplicados_rows.append([
                            device_id, ingreso_id or "", "mg_manual", mg_norm, "", alquiler_a_manual or "",
                            dprev[0], dprev[1], alquilado_ing_prev, alquiler_a_ing_prev,
                        ])
                        applied_count += 1

                # MG no alquilados (excluyendo ubicaciones)
                cur.execute(
                    r"""
                    SELECT d.id, d.numero_interno, d.numero_serie,
                           COALESCE(l.nombre,'') AS ubicacion,
                           COALESCE(c.razon_social,'') AS cliente
                      FROM devices d
                      JOIN customers c ON c.id = d.customer_id
                      LEFT JOIN LATERAL (
                        SELECT t.ubicacion_id
                          FROM ingresos t
                         WHERE t.device_id = d.id
                         ORDER BY COALESCE(t.fecha_ingreso, t.fecha_creacion) DESC, t.id DESC
                         LIMIT 1
                      ) last ON TRUE
                      LEFT JOIN locations l ON l.id = last.ubicacion_id
                     WHERE (
                            (d.numero_interno IS NOT NULL AND d.numero_interno ~ '^MG\\s\\d{4}$')
                         OR (d.numero_serie IS NOT NULL AND d.numero_serie ~ '^MG\\s\\d{4}$')
                           )
                    """
                )
                rows = cur.fetchall() or []
                for rid, mg, ns, ubic, cliente_act in rows:
                    mg_norm = self._norm_mg(mg) or (self._norm_mg(ns) if ns else None)
                    if mg_norm and mg_norm in mg_alquilados_aplicados:
                        continue
                    # Excluir ubicaciones definidas
                    if self._is_excluded_location(ubic):
                        continue
                    mg_no_alquilados_rows.append([
                        rid, mg_norm or (ns or ""), ns or "", (ubic or ""), (cliente_act or ""),
                    ])

                # Si es dry-run, no dejamos la transacciÃ³n abierta
                if dry:
                    transaction.set_rollback(True)

        # Aplicar renombrados de cÃ³digos en devices.n_de_control (FROM:TO)
        rename_list = opts.get("rename_code") or []
        if rename_list:
            with transaction.atomic():
                with connection.cursor() as cur:
                    for item in rename_list:
                        try:
                            from_code, to_code = [x.strip() for x in str(item).split(":", 1)]
                        except Exception:
                            conflictos_rows.append([
                                "rename", "", item, "", "code", "", 0, "Formato invÃ¡lido (esperado FROM:TO)"
                            ])
                            continue
                        # Conteo previo
                        cur.execute("SELECT COUNT(*) FROM devices WHERE numero_interno=%s", [from_code])
                        cnt = int(cur.fetchone()[0] or 0)
                        if not dry and cnt:
                            cur.execute(
                                "UPDATE devices SET numero_interno=%s WHERE numero_interno=%s",
                                [to_code, from_code],
                            )
                        aplicados_rows.append([
                            "-", "-", "rename_code", f"{from_code} -> {to_code}", "", "",
                            "-", "-", "-", f"rows={cnt}",
                        ])

        # Escribir backups y reportes
        # Backups
        self._write_backup_csv(
            os.path.join(docs_dir, "backup_devices_alquiler.csv"),
            backup_devices[0], backup_devices[1:]
        )
        self._write_backup_csv(
            os.path.join(docs_dir, "backup_ingresos_ultimo_alquiler.csv"),
            backup_ingresos[0], backup_ingresos[1:]
        )

        # Reporte principal
        result_path = os.path.join(docs_dir, "alquileres_sync_result.xlsx")
        self._write_result_xlsx(
            result_path,
            {
                "aplicados": aplicados_rows,
                "conflictos": conflictos_rows,
                "no_encontrados": no_encontrados_rows,
                "mg_no_alquilados": mg_no_alquilados_rows,
            },
        )

        # Archivo aparte con detalle de conflictos
        conflictos_path = os.path.join(docs_dir, "alquileres_conflictos_detalle.xlsx")
        self._write_result_xlsx(
            conflictos_path,
            {
                "conflictos_detalle": conflictos_detalle_rows,
            },
        )

        self.stdout.write(
            ("DRY-RUN " if dry else "APLICADO ") +
            f"OK. Reporte: {result_path} | Conflictos: {conflictos_path} | Backups en {docs_dir}"
        )
        return
