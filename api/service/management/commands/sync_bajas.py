from django.core.management.base import BaseCommand
from django.db import connection, transaction
from typing import Dict, List, Tuple, Optional, Any
import os
import csv
import re


class Command(BaseCommand):
    help = (
        "Procesa 'Copia de HISTORICO DE BAJAS 2024.xlsx' (col B estado, G código interno, J fecha) y: "
        "BAJA → setea última ubicación a 'Desguace'; ALTA no alquilados → los lista en reporte."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Simula cambios y genera reportes sin escribir en BD (default)",
        )
        parser.add_argument(
            "--apply",
            action="store_true",
            help="Aplica cambios en una transacción atómica",
        )
        parser.add_argument(
            "--docs-dir",
            default=None,
            help="Directorio de salida para reportes (por defecto autodetecta 'docs' o '../docs')",
        )
        parser.add_argument(
            "--file",
            default=None,
            help="Ruta alternativa del Excel de bajas/altas (default: docs/Copia de HISTORICO DE BAJAS 2024.xlsx)",
        )
        parser.add_argument(
            "--sheet",
            default=None,
            help="Nombre de hoja a leer (default: primera hoja)",
        )

    # ---- Utilidades de normalización ----
    @staticmethod
    def _norm_text(s: Optional[str]) -> str:
        try:
            import unicodedata
            s2 = (s or "").strip()
            s2 = "".join(
                c for c in unicodedata.normalize("NFKD", s2)
                if not unicodedata.combining(c)
            )
            return " ".join(s2.split())
        except Exception:
            return (s or "").strip()

    @staticmethod
    def _norm_lower_nodiac(s: Optional[str]) -> str:
        try:
            import unicodedata
            s2 = (s or "").strip().lower()
            s2 = "".join(
                c for c in unicodedata.normalize("NFKD", s2)
                if not unicodedata.combining(c)
            )
            return " ".join(s2.split())
        except Exception:
            return (s or "").strip().lower()

    @staticmethod
    def _norm_ns(ns: Optional[str]) -> str:
        s = (ns or "").strip()
        s = s.upper()
        s = s.replace(" ", "").replace("-", "")
        return s

    @staticmethod
    def _norm_stockcode(code: Optional[str]) -> Optional[str]:
        """Normaliza códigos de stock (MG|NM|NV|CE) a 'XX ####'."""
        if not code:
            return None
        s = (code or "").upper().strip()
        m = re.match(r"^(MG|NM|NV|CE)\s*(\d{1,4})$", s)
        if not m:
            m2 = re.match(r"^(MG|NM|NV|CE)[^0-9]*(\d{1,4})$", s)
            if not m2:
                return None
            pref, num = m2.group(1), m2.group(2)
        else:
            pref, num = m.group(1), m.group(2)
        return f"{pref} {num.zfill(4)}"

    @staticmethod
    def _ensure_docs_dir(path: str) -> None:
        try:
            os.makedirs(path, exist_ok=True)
        except Exception:
            pass

    @staticmethod
    def _fmt_dt(x: Any) -> str:
        try:
            if x is None:
                return ""
            if hasattr(x, "tzinfo"):
                try:
                    x = x.replace(tzinfo=None)
                except Exception:
                    return str(x)
            try:
                return x.strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                return str(x)
        except Exception:
            return ""

    # ---- Lectura Excel ----
    def _pick_docs_dir(self, user_docs_dir: Optional[str]) -> str:
        if user_docs_dir:
            return user_docs_dir
        if os.path.isdir(os.path.join("docs")):
            return os.path.join("docs")
        if os.path.isdir(os.path.join("..", "docs")):
            return os.path.join("..", "docs")
        return os.path.join("docs")

    def _resolve_input(self, base_docs: str, filename: str) -> str:
        p1 = os.path.join(base_docs, filename)
        if os.path.exists(p1):
            return p1
        alt = os.path.join("..", base_docs, filename) if not base_docs.startswith("..") else os.path.join(base_docs[3:], filename)
        return alt if os.path.exists(alt) else p1

    def _load_bajas(self, xlsx_path: str, sheet_name: Optional[str]) -> List[Dict[str, Any]]:
        from openpyxl import load_workbook
        rows: List[Dict[str, Any]] = []
        try:
            wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        except Exception as e:
            self.stderr.write(f"No se pudo abrir '{xlsx_path}': {e}")
            return rows
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb[wb.sheetnames[0]]
        # Col B (1), G (6), J (9) 0-based
        for row in ws.iter_rows(min_row=1):
            try:
                estado = row[1].value if len(row) >= 2 else None
                code = row[6].value if len(row) >= 7 else None
                fecha_baja = row[9].value if len(row) >= 10 else None
            except Exception:
                estado = None
                code = None
                fecha_baja = None
            estado_s = (str(estado).strip().upper() if estado is not None else "")
            code_s = (str(code).strip() if code is not None else "")
            if not estado_s and not code_s:
                continue
            # Saltar encabezados obvios
            if estado_s in ("ESTADO",) or self._norm_lower_nodiac(code_s).startswith("numero interno"):
                continue
            rows.append({
                "estado": estado_s,
                "code_raw": code_s,
                "fecha_baja": fecha_baja,
                "source": os.path.basename(xlsx_path),
            })
        return rows

    # ---- BD helpers ----
    def _fetch_device_candidates(self, cur, code_norm: str) -> List[Dict[str, Any]]:
        cur.execute(
            """
            SELECT id, customer_id, numero_interno, n_de_control, numero_serie, alquilado, alquiler_a
              FROM devices
             WHERE (numero_interno = %s)
                OR (numero_serie = %s)
                OR (REPLACE(REPLACE(UPPER(numero_serie),' ',''),'-','') = %s)
            """,
            [code_norm, code_norm, self._norm_ns(code_norm)],
        )
        return [
            {"id": r[0], "customer_id": r[1], "numero_interno": r[2], "n_de_control": r[3], "numero_serie": r[4], "alquilado": r[5], "alquiler_a": r[6]}
            for r in (cur.fetchall() or [])
        ]

    def _fetch_last_ingreso_id(self, cur, device_id: int) -> Optional[int]:
        cur.execute(
            """
            SELECT id
              FROM ingresos
             WHERE device_id = %s
             ORDER BY COALESCE(fecha_ingreso, fecha_creacion) DESC, id DESC
             LIMIT 1
            """,
            [device_id],
        )
        row = cur.fetchone()
        return int(row[0]) if row else None

    def _fetch_last_ingreso_row(self, cur, device_id: int) -> Optional[Dict[str, Any]]:
        cur.execute(
            """
            SELECT t.id, t.estado, t.fecha_ingreso, t.fecha_creacion,
                   l.id AS ubic_id, COALESCE(l.nombre,'') AS ubic_nombre,
                   t.alquilado, t.alquiler_a
              FROM ingresos t
              LEFT JOIN locations l ON l.id = t.ubicacion_id
             WHERE t.device_id = %s
             ORDER BY COALESCE(t.fecha_ingreso, t.fecha_creacion) DESC, t.id DESC
             LIMIT 1
            """,
            [device_id],
        )
        r = cur.fetchone()
        if not r:
            return None
        return {
            "id": r[0], "estado": r[1], "fecha_ingreso": r[2], "fecha_creacion": r[3],
            "ubic_id": r[4], "ubic_nombre": r[5], "alquilado": r[6], "alquiler_a": r[7],
        }

    def _ensure_location_id(self, cur, target_name: str) -> int:
        cur.execute(
            "SELECT id FROM locations WHERE LOWER(nombre)=LOWER(%s) LIMIT 1",
            [target_name],
        )
        r = cur.fetchone()
        if r:
            return int(r[0])
        # Buscar por normalización (sin acentos)
        try:
            cur.execute("SELECT id, nombre FROM locations")
            rows = cur.fetchall() or []
            tgt = self._norm_lower_nodiac(target_name)
            for lid, lname in rows:
                if self._norm_lower_nodiac(lname) == tgt:
                    return int(lid)
        except Exception:
            pass
        # Crear si no existe
        cur.execute("INSERT INTO locations(nombre) VALUES (%s) RETURNING id", [target_name])
        return int(cur.fetchone()[0])

    def _pick_canonical_device(self, devices: List[Dict[str, Any]], code_norm: str) -> Tuple[Optional[Dict[str, Any]], bool]:
        if not devices:
            return None, False
        def dev_code(d: Dict[str, Any]) -> Optional[str]:
            # Preferir numero_interno; si no, ver si numero_serie ya es el code
            nin = (d.get("numero_interno") or "").strip()
            if nin:
                return (nin or None)
            ns = (d.get("numero_serie") or "").strip()
            return ns or None
        keys = [dev_code(d) or "" for d in devices]
        unique_keys = set([k for k in keys if k])
        canonical = max(devices, key=lambda d: int(d.get("id")))
        is_conflict = len(unique_keys) > 1 and (code_norm not in unique_keys)
        return canonical, is_conflict

    # ---- Reportes ----
    def _write_backup_csv(self, path: str, headers: List[str], rows: List[List[Any]]) -> None:
        try:
            with open(path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(headers)
                for r in rows:
                    w.writerow(r)
        except Exception as e:
            self.stderr.write(f"No se pudo escribir backup '{path}': {e}")

    def _write_result_xlsx(self, path: str, tabs: Dict[str, List[List[Any]]]) -> None:
        try:
            from openpyxl import Workbook
            wb = Workbook()
            first_name = next(iter(tabs.keys())) if tabs else "baja_aplicados"
            ws0 = wb.active
            ws0.title = (first_name[:31] or "sheet")
            for row in tabs.get(first_name, []):
                ws0.append(row)
            for name, rows in tabs.items():
                if name == first_name:
                    continue
                ws = wb.create_sheet(title=name[:31] or "sheet")
                for r in rows:
                    ws.append(r)
            wb.save(path)
        except Exception as e:
            self.stderr.write(f"No se pudo escribir reporte '{path}': {e}")

    def handle(self, *args, **opts):
        dry = True
        if opts.get("apply"):
            dry = False
        if opts.get("dry_run"):
            dry = True

        docs_dir = self._pick_docs_dir(opts.get("docs_dir"))
        self._ensure_docs_dir(docs_dir)

        bajas_path = opts.get("file") or self._resolve_input(docs_dir, "Copia de HISTORICO DE BAJAS 2024.xlsx")
        sheet = opts.get("sheet")

        rows_in = self._load_bajas(bajas_path, sheet)
        if not rows_in:
            self.stderr.write("No se encontraron filas en el Excel de bajas/altas. Abortando.")
            return

        baja_aplicados: List[List[Any]] = [[
            "device_id", "ingreso_id", "codigo", "ns_device", "ubicacion_anterior", "ubicacion_nueva", "fecha_baja", "source",
        ]]
        alta_no_alquilados: List[List[Any]] = [[
            "device_id", "codigo", "ns_device", "ultima_ubicacion", "cliente", "alquilado_device", "alquilado_ingreso", "last_fecha", "source",
        ]]
        alta_marcados: List[List[Any]] = [[
            "device_id", "ingreso_id", "codigo", "ns_device", "alquilado_device_prev", "alquilado_ingreso_prev", "source",
        ]]
        no_encontrados: List[List[Any]] = [["source", "estado", "codigo_raw", "codigo_norm", "fecha_baja"]]
        conflictos: List[List[Any]] = [["source", "estado", "codigo_raw", "codigo_norm", "coincidencias", "detalle_candidatos"]]
        conflictos_detalle: List[List[Any]] = [[
            "source", "estado", "codigo_raw", "codigo_norm", "fecha_baja",
            "device_id", "customer_id", "cliente_bd", "marca", "modelo", "tipo_equipo",
            "numero_interno", "numero_serie", "alquilado_device", "alquiler_a_device",
            "last_ingreso_id", "last_estado", "last_fecha_ingreso", "last_fecha_creacion",
            "last_ubicacion_id", "last_ubicacion_nombre", "alquilado_ingreso", "alquiler_a_ingreso",
        ]]

        backups_ingresos: List[List[Any]] = [["ingreso_id", "ubicacion_id_prev", "ubicacion_nombre_prev", "comentarios_prev"]]

        with transaction.atomic():
            with connection.cursor() as cur:
                loc_desguace_id = self._ensure_location_id(cur, "Desguace")
                for r in rows_in:
                    estado = (r.get("estado") or "").upper().strip()
                    code_raw = r.get("code_raw") or ""
                    code_norm = self._norm_stockcode(code_raw)
                    fecha_baja = r.get("fecha_baja")
                    source = r.get("source")

                    if not code_norm:
                        no_encontrados.append([source, estado, code_raw, "", self._fmt_dt(fecha_baja)])
                        continue

                    devs = self._fetch_device_candidates(cur, code_norm)
                    if not devs:
                        no_encontrados.append([source, estado, code_raw, code_norm, self._fmt_dt(fecha_baja)])
                        continue

                    canonical, is_conflict = self._pick_canonical_device(devs, code_norm)
                    if is_conflict:
                        cand_lines: List[str] = []
                        for d in devs:
                            did = int(d.get("id"))
                            cur.execute(
                                """
                                SELECT d.numero_interno, d.numero_serie,
                                       COALESCE(c.razon_social,'') AS cliente,
                                       c.id AS customer_id,
                                       COALESCE(b.nombre,'') AS marca,
                                       COALESCE(m.nombre,'') AS modelo,
                                       COALESCE(m.tipo_equipo,'') AS tipo_equipo,
                                       d.alquilado, d.alquiler_a
                                  FROM devices d
                                  JOIN customers c ON c.id = d.customer_id
                                  LEFT JOIN marcas b ON b.id = d.marca_id
                                  LEFT JOIN models m ON m.id = d.model_id
                                 WHERE d.id=%s
                                """,
                                [did],
                            )
                            dr = cur.fetchone() or [None, None, "", None, "", "", "", None, None]
                            li = self._fetch_last_ingreso_row(cur, did) or {}
                            ubic = li.get("ubic_nombre") or ""
                            cand_lines.append(
                                f"id={did}|MG/NRO={dr[0] or ''}|NS={dr[1] or ''}|Cliente={dr[2] or ''}|Ubic={ubic}"
                            )
                            conflictos_detalle.append([
                                source, estado, code_raw, code_norm, self._fmt_dt(fecha_baja),
                                did, dr[3], dr[2], dr[4], dr[5], dr[6],
                                dr[0], dr[1], dr[7], dr[8],
                                li.get("id"), li.get("estado"), self._fmt_dt(li.get("fecha_ingreso")), self._fmt_dt(li.get("fecha_creacion")),
                                li.get("ubic_id"), li.get("ubic_nombre"), li.get("alquilado"), li.get("alquiler_a"),
                            ])
                        conflictos.append([source, estado, code_raw, code_norm, len(devs), " ; ".join(cand_lines)])
                        continue

                    dev = canonical or devs[0]
                    device_id = int(dev.get("id"))
                    last = self._fetch_last_ingreso_row(cur, device_id)

                    # Si no hay ingresos, solo reportar
                    if not last:
                        no_encontrados.append([source, estado, code_raw, code_norm, self._fmt_dt(fecha_baja)])
                        continue

                    # ALTA: si está alquilado, marcar en SE; si no, listar con última fecha
                    if estado == "ALTA":
                        dev_alq_prev = bool(dev.get("alquilado") or False)
                        last_alq_prev = bool(last.get("alquilado") or False)
                        if dev_alq_prev or last_alq_prev:
                            if not dry:
                                # Asegurar flags de alquiler en device y último ingreso
                                cur.execute(
                                    "UPDATE devices SET alquilado=true WHERE id=%s",
                                    [device_id],
                                )
                                cur.execute(
                                    "UPDATE ingresos SET alquilado=true WHERE id=%s",
                                    [last.get("id")],
                                )
                            alta_marcados.append([
                                device_id, last.get("id"), code_norm, dev.get("numero_serie") or "",
                                dev_alq_prev, last_alq_prev, source,
                            ])
                        else:
                            # Obtener cliente para reporte
                            cur.execute("SELECT COALESCE(c.razon_social,'') FROM customers c JOIN devices d ON d.customer_id=c.id WHERE d.id=%s", [device_id])
                            cliente_name = (cur.fetchone() or [""])[0]
                            # Última fecha de movimiento (prefiere fecha_ingreso)
                            last_fecha = last.get("fecha_ingreso") or last.get("fecha_creacion")
                            alta_no_alquilados.append([
                                device_id, code_norm, dev.get("numero_serie") or "",
                                last.get("ubic_nombre") or "", cliente_name,
                                dev_alq_prev, last_alq_prev, self._fmt_dt(last_fecha), source,
                            ])
                        continue

                    # BAJA: mover a Desguace en último ingreso
                    if estado == "BAJA":
                        # Obtener cliente para incluir en reporte
                        cur.execute("SELECT COALESCE(c.razon_social,'') FROM customers c JOIN devices d ON d.customer_id=c.id WHERE d.id=%s", [device_id])
                        cliente_name = (cur.fetchone() or [""])[0]
                        # Backup previo
                        backups_ingresos.append([
                            last.get("id"), last.get("ubic_id"), last.get("ubic_nombre"), None,
                        ])
                        if not dry:
                            cur.execute(
                                "UPDATE ingresos SET ubicacion_id=%s WHERE id=%s",
                                [loc_desguace_id, last.get("id")],
                            )
                        baja_aplicados.append([
                            device_id, last.get("id"), code_norm, dev.get("numero_serie") or "",
                            last.get("ubic_nombre") or "", "Desguace", self._fmt_dt(fecha_baja), source,
                        ])
                        continue

                    # Otros estados: ignorar o reportar
                    conflictos.append([source, estado, code_raw, code_norm, 1, f"Estado no soportado: {estado}"])

                if dry:
                    transaction.set_rollback(True)

        # Escribir reportes
        result_path = os.path.join(docs_dir, "bajas_sync_result.xlsx")
        self._write_result_xlsx(
            result_path,
            {
                "baja_aplicados": baja_aplicados,
                "alta_marcados": alta_marcados,
                "alta_no_alquilados": alta_no_alquilados,
                "no_encontrados": no_encontrados,
                "conflictos": conflictos,
            },
        )
        det_path = os.path.join(docs_dir, "bajas_conflictos_detalle.xlsx")
        self._write_result_xlsx(
            det_path,
            {"conflictos_detalle": conflictos_detalle},
        )
        # Backups
        self._write_backup_csv(
            os.path.join(docs_dir, "backup_ingresos_bajas.csv"),
            backups_ingresos[0], backups_ingresos[1:]
        )

        self.stdout.write(
            ("DRY-RUN " if dry else "APLICADO ") +
            f"OK. Reportes: {result_path} | {det_path}"
        )
