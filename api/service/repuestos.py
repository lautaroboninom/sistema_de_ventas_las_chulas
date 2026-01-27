import datetime as dt
import hashlib
import os
import re
from decimal import Decimal, InvalidOperation
from typing import Dict, List, Tuple

from django.conf import settings
from django.db import connection
from django.utils import timezone


def _is_header_unificados(code: str, name: str, proveedor_raw) -> bool:
    h = f"{code} {name}".strip().lower()
    if not h:
        return False
    if "codigo" in h or "código" in h:
        return True
    if "descripcion" in h or "descripción" in h:
        return True
    if proveedor_raw is None:
        return False
    p = str(proveedor_raw).strip().lower()
    if "proveedor" in p:
        return True
    return False


def load_repuestos_unificados_from_excel(path: str | None, sheet: str | None = None):
    """
    Layout esperado:
      - Col A: codigo (principal)
      - Col B: descripcion / nombre
      - Col E: proveedor (texto)
    """
    if not path:
        return {}, []
    if not os.path.exists(path):
        return {}, [{"error": "file_not_found", "path": path}]

    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    sheet_names = [sheet] if sheet else wb.sheetnames

    items = {}
    conflicts = []

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            code_raw = row[0] if len(row) > 0 else None  # A
            name_raw = row[1] if len(row) > 1 else None  # B
            prov_raw = row[4] if len(row) > 4 else None  # E

            code = _norm_code(code_raw)
            name = _norm_name(name_raw)
            prov = _norm_name(prov_raw)

            if row_idx <= 3 and _is_header_unificados(code, name, prov_raw):
                continue
            if not code or not name:
                continue

            prev = items.get(code)
            if prev and prev.get("nombre") != name:
                conflicts.append(
                    {"codigo": code, "nombre_prev": prev.get("nombre"), "nombre_new": name, "sheet": sheet_name, "row": row_idx}
                )

            items[code] = {"codigo": code, "nombre": name, "proveedor": prov}

    return items, conflicts


def sync_catalogo_repuestos_unificados(path: str | None = None, sheet: str | None = None, deactivate_missing: bool = False):
    """
    - Upsert en catalogo_repuestos (sin tocar costos)
    - Crea proveedores_externos si no existen (por nombre, case-insensitive)
    - Crea relación repuestos_proveedores (prioridad=1) si no existe
    - Por defecto NO desactiva faltantes
    """
    file_path = path or getattr(settings, "REPUESTOS_UNIFICADOS_FILE", None)
    items, conflicts = load_repuestos_unificados_from_excel(file_path, sheet=sheet)
    if not items:
        return {"count": 0, "conflicts": conflicts}

    codes = list(items.keys())
    proveedores = sorted({(it.get("proveedor") or "").strip() for it in items.values() if (it.get("proveedor") or "").strip()})

    with connection.cursor() as cur:
        # 1) Upsert catálogo (NO toca costos)
        rows = [(it["codigo"], it["nombre"]) for it in items.values()]
        cur.executemany(
            """
            INSERT INTO catalogo_repuestos (codigo, nombre, activo, updated_at)
            VALUES (%s,%s,TRUE,NOW())
            ON CONFLICT (codigo) DO UPDATE SET
              nombre=EXCLUDED.nombre,
              activo=TRUE,
              updated_at=NOW()
            """,
            rows,
        )

        # 2) Crear proveedores faltantes
        proveedor_id_by_lower = {}
        if proveedores:
            lowers = [p.lower() for p in proveedores]
            cur.execute("SELECT id, nombre FROM proveedores_externos WHERE LOWER(nombre)=ANY(%s)", [lowers])
            for pid, nombre in (cur.fetchall() or []):
                if nombre:
                    proveedor_id_by_lower[str(nombre).lower()] = pid

            for p in proveedores:
                pl = p.lower()
                if pl in proveedor_id_by_lower:
                    continue
                cur.execute("INSERT INTO proveedores_externos (nombre) VALUES (%s)", [p])
                cur.execute(
                    "SELECT id FROM proveedores_externos WHERE LOWER(nombre)=LOWER(%s) ORDER BY id DESC LIMIT 1",
                    [p],
                )
                row = cur.fetchone()
                if row:
                    proveedor_id_by_lower[pl] = row[0]

        # 3) Asociar repuestos_proveedores
        if proveedor_id_by_lower:
            cur.execute("SELECT id, codigo FROM catalogo_repuestos WHERE codigo=ANY(%s)", [codes])
            repuesto_id_by_code = {str(c).upper(): rid for (rid, c) in (cur.fetchall() or [])}

            links = []
            for it in items.values():
                p = (it.get("proveedor") or "").strip()
                if not p:
                    continue
                rid = repuesto_id_by_code.get(it["codigo"].upper())
                pid = proveedor_id_by_lower.get(p.lower())
                if rid and pid:
                    links.append((rid, pid))

            if links:
                cur.executemany(
                    """
                    INSERT INTO repuestos_proveedores (repuesto_id, proveedor_id, prioridad, created_at, updated_at)
                    VALUES (%s,%s,1,NOW(),NOW())
                    ON CONFLICT (repuesto_id, proveedor_id) DO NOTHING
                    """,
                    links,
                )

        # 4) Opcional: desactivar los que no están en el Excel
        if deactivate_missing:
            cur.execute("UPDATE catalogo_repuestos SET activo=FALSE WHERE codigo <> ALL(%s)", [codes])

    return {"count": len(items), "conflicts": conflicts}


def _norm_code(val) -> str:
    if val is None:
        return ""
    if isinstance(val, float):
        if val.is_integer():
            return str(int(val))
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.strip().upper()


def _norm_name(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _parse_cost(val) -> Decimal | None:
    if val is None:
        return None
    if isinstance(val, Decimal):
        return val
    if isinstance(val, (int, float)):
        return Decimal(str(val))
    s = str(val).strip()
    if not s:
        return None
    s = s.replace("$", "").replace("ARS", "").replace(" ", "")
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    else:
        s = s.replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _to_decimal(val) -> Decimal | None:
    if val is None:
        return None
    if isinstance(val, Decimal):
        return val
    if isinstance(val, (int, float)):
        return Decimal(str(val))
    s = str(val).strip()
    if not s:
        return None
    s = s.replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def get_repuestos_config() -> Dict:
    with connection.cursor() as cur:
        cur.execute(
            """
            SELECT id, dolar_ars, multiplicador_general
            FROM repuestos_config
            ORDER BY id
            LIMIT 1
            """
        )
        row = cur.fetchone()
        if not row:
            cur.execute(
                """
                INSERT INTO repuestos_config (dolar_ars, multiplicador_general)
                VALUES (0, 1)
                RETURNING id, dolar_ars, multiplicador_general
                """
            )
            row = cur.fetchone()
    return {"id": row[0], "dolar_ars": row[1], "multiplicador_general": row[2]}


def calc_costo_ars(costo_usd, dolar_ars):
    costo = _to_decimal(costo_usd)
    dolar = _to_decimal(dolar_ars)
    if costo is None or dolar is None:
        return None
    return (costo * dolar).quantize(Decimal("0.01"))


def calc_precio_venta(costo_usd, dolar_ars, multiplicador_general, multiplicador_individual):
    costo = _to_decimal(costo_usd)
    dolar = _to_decimal(dolar_ars)
    mult_gen = _to_decimal(multiplicador_general)
    mult_ind = _to_decimal(multiplicador_individual)
    if costo is None or dolar is None:
        return None
    mult = mult_ind if mult_ind is not None else mult_gen
    if mult is None:
        return None
    return (costo * dolar * mult).quantize(Decimal("0.01"))


def _fallback_code_from_name(name: str) -> str:
    # Stable, ASCII-only code for rows without explicit code.
    raw = (name or "").strip().upper()
    base = re.sub(r"[^A-Z0-9]+", "-", raw).strip("-")
    if not base:
        base = "SIN-NOMBRE"
    base = base[:40].strip("-")
    h = hashlib.sha1((name or "").encode("utf-8")).hexdigest()[:8].upper()
    return f"NC-{base}-{h}"


def _is_header(code: str, name: str, cost) -> bool:
    c = (code or "").lower()
    n = (name or "").lower()
    if "codigo" in c:
        return True
    if "repuesto" in n or "descripcion" in n:
        return True
    if isinstance(cost, str) and ("costo" in cost.lower() or "precio" in cost.lower()):
        return True
    return False


def load_repuestos_from_excel(path: str, sheet: str | None = None) -> Tuple[Dict[str, Dict], List[Dict]]:
    from openpyxl import load_workbook

    if not path:
        raise FileNotFoundError("REPUESTOS_COSTOS_FILE no definido")
    if not os.path.exists(path):
        raise FileNotFoundError(f"No existe archivo: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = [sheet] if sheet else wb.sheetnames
    items: Dict[str, Dict] = {}
    conflicts: List[Dict] = []

    for sheet_name in sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            code_raw = row[1] if len(row) > 1 else None  # col B
            cost_raw = row[3] if len(row) > 3 else None  # col D
            name_raw = row[5] if len(row) > 5 else None  # col F

            code = _norm_code(code_raw)
            name = _norm_name(name_raw)
            cost = _parse_cost(cost_raw)

            if not name:
                continue
            if row_idx <= 2 and _is_header(code, name, cost_raw):
                continue
            if cost is None:
                continue
            if not code:
                code = _fallback_code_from_name(name)

            cost = cost.quantize(Decimal("0.01"))
            prev = items.get(code)
            if prev and (
                prev["nombre"] != name
                or prev["costo_usd"] != cost
            ):
                conflicts.append({
                    "codigo": code,
                    "nombre_prev": prev["nombre"],
                    "costo_prev": str(prev["costo_usd"]),
                    "nombre_new": name,
                    "costo_new": str(cost),
                    "sheet": sheet_name,
                    "row": row_idx,
                })
            items[code] = {"codigo": code, "nombre": name, "costo_usd": cost}

    return items, conflicts


def sync_catalogo_repuestos(path: str | None = None, sheet: str | None = None, deactivate_missing: bool = True) -> Dict:
    file_path = path or getattr(settings, "REPUESTOS_COSTOS_FILE", None)
    items, conflicts = load_repuestos_from_excel(file_path, sheet=sheet)
    if not items:
        return {"count": 0, "conflicts": conflicts}

    cfg = get_repuestos_config()
    dolar_ars = _to_decimal(cfg.get("dolar_ars")) or Decimal("0")
    mtime = os.path.getmtime(file_path)
    mtime_dt = timezone.make_aware(dt.datetime.fromtimestamp(mtime), timezone.get_current_timezone())
    rows = []
    for it in items.values():
        costo_usd = it["costo_usd"]
        costo_ars = (costo_usd * dolar_ars).quantize(Decimal("0.01")) if dolar_ars else Decimal("0.00")
        rows.append((it["codigo"], it["nombre"], costo_usd, costo_ars, mtime_dt))

    with connection.cursor() as cur:
        cur.executemany(
            """
            INSERT INTO catalogo_repuestos (codigo, nombre, costo_usd, costo_neto, activo, source_mtime, updated_at)
            VALUES (%s,%s,%s,%s,TRUE,%s,NOW())
            ON CONFLICT (codigo) DO UPDATE SET
              nombre=EXCLUDED.nombre,
              costo_usd=EXCLUDED.costo_usd,
              costo_neto=EXCLUDED.costo_neto,
              activo=TRUE,
              source_mtime=EXCLUDED.source_mtime,
              updated_at=NOW()
            """,
            rows,
        )
        if deactivate_missing:
            cur.execute(
                "UPDATE catalogo_repuestos SET activo=FALSE WHERE source_mtime IS DISTINCT FROM %s",
                [mtime_dt],
            )
        if dolar_ars:
            cur.execute(
                """
                UPDATE quote_items qi
                   SET costo_u_neto = ROUND(cr.costo_usd * %s, 2)
                  FROM catalogo_repuestos cr
                 WHERE qi.tipo='repuesto'
                   AND qi.costo_u_neto IS NULL
                   AND qi.repuesto_codigo IS NOT NULL
                   AND UPPER(qi.repuesto_codigo)=UPPER(cr.codigo)
                   AND cr.activo
                   AND cr.costo_usd IS NOT NULL
                """,
                [dolar_ars],
            )

    return {"count": len(rows), "conflicts": conflicts}
