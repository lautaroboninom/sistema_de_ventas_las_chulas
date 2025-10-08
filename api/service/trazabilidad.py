import os
import datetime as _dt
from typing import Optional, Tuple, Dict, Any

from django.conf import settings


def _norm_txt(s: str) -> str:
    import unicodedata, re
    s = str(s or "")
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def _norm_serial(s: str) -> str:
    import re
    s = str(s or "").strip().upper()
    s = re.sub(r"[\s\-_/]", "", s)
    return s


def _norm_file_token(name: str) -> str:
    base = os.path.splitext(os.path.basename(name))[0]
    return _norm_txt(base)


def list_excel_files(brand_hint: Optional[str] = None) -> list:
    root = getattr(settings, "TRAZABILIDAD_ROOT", None)
    if not root or not os.path.isdir(root):
        return []
    exts = {".xlsx", ".xls"}
    out = []
    for base, _dirs, files in os.walk(root):
        for fn in files:
            ext = os.path.splitext(fn)[1].lower()
            if ext in exts:
                # Todos los archivos comienzan con '@'
                if not fn.startswith("@"):  # filtrar por convención dada
                    continue
                out.append(os.path.join(base, fn))
    if brand_hint:
        bh = _norm_txt(brand_hint)
        # priorizar los que contengan el nombre de la marca
        def score(p: str) -> int:
            t = _norm_file_token(p)
            return 1 if (bh and (bh == t or bh in t or t in bh)) else 0
        out.sort(key=score, reverse=True)
        # si hay coincidencias (score 1), podemos limitar la búsqueda a ellos
        top = [p for p in out if score(p) == 1]
        if top:
            return top
    return out


def _parse_date(val) -> Optional[_dt.date]:
    if not val:
        return None
    if isinstance(val, _dt.datetime):
        return val.date()
    if isinstance(val, _dt.date):
        return val
    # Excel serial (heurística)
    if isinstance(val, (int, float)):
        # Rango típico de fechas Excel serial (~1950-2100)
        if 20000 <= float(val) <= 60000:
            try:
                base = _dt.date(1899, 12, 30)  # convención Excel (Windows)
                return base + _dt.timedelta(days=float(val))
            except Exception:
                pass
    s = str(val).strip()
    if not s:
        return None
    # Try several formats
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%m/%d/%Y", "%d.%m.%Y"):
        try:
            return _dt.datetime.strptime(s, fmt).date()
        except Exception:
            pass
    # Fallback: try fromisoformat tolerant
    try:
        return _dt.date.fromisoformat(s)
    except Exception:
        return None


def _iter_rows_from_xlsx(path: str):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Buscar fila de encabezados en las primeras 50 filas
        header_idx = None
        hdrs = None
        scan_limit = 50
        rgen = ws.iter_rows(values_only=True)
        buffered = []
        for idx in range(scan_limit):
            try:
                row = next(rgen)
            except StopIteration:
                break
            buffered.append(row)
            cand = [(_norm_txt(h) if h is not None else "") for h in row]
            # ¿Contiene al menos la columna de serie preferida o un hint y alguna fecha?
            s_col_tmp = _pick_serial_col(cand)
            d_cols_tmp = _pick_date_cols(cand)
            if s_col_tmp is not None and d_cols_tmp:
                header_idx = idx
                hdrs = cand
                break
        if header_idx is None:
            continue
        # Generar iterador de filas de datos desde header_idx+1
        def _rows_iter():
            # Consumir las filas que ya bufferizamos por delante del header
            for r in buffered[header_idx+1:]:
                yield r
            # Continuar con el resto de filas del worksheet
            for r in rgen:
                yield r
            # Y si quedaran filas que no estaban en el primer iterador (por seguridad), reabrir iter_rows saltando
            # Nota: en modo read_only el iterador es único; esto cubre casos donde header está muy arriba
        yield sheet_name, hdrs, _rows_iter()


def _iter_rows_from_xls(path: str):
    import xlrd
    book = xlrd.open_workbook(path)
    for sheet in book.sheets():
        if sheet.nrows == 0:
            continue
        header_idx = None
        hdrs = None
        for r in range(min(sheet.nrows, 50)):
            headers = [sheet.cell_value(r, c) for c in range(sheet.ncols)]
            cand = [(_norm_txt(h) if h is not None else "") for h in headers]
            s_col_tmp = _pick_serial_col(cand)
            d_cols_tmp = _pick_date_cols(cand)
            if s_col_tmp is not None and d_cols_tmp:
                header_idx = r
                hdrs = cand
                break
        if header_idx is None:
            continue
        def _row_iter():
            for r in range(header_idx + 1, sheet.nrows):
                yield [sheet.cell_value(r, c) for c in range(sheet.ncols)]
        yield sheet.name, hdrs, _row_iter()


SERIAL_HINTS = (
    # Preferido (según tu estructura): "Ítem - Artículo - Partida - Cód."
    "item_articulo_partida_cod",
    # Otras variantes comunes (fallback heurístico)
    "numero_de_serie", "nro_de_serie", "n_de_serie", "n_serie",
    "nro_serie", "serie", "serial", "num_serie", "num_de_serie",
    "nro", "ns", "n_s"
)

DATE_HINTS_WEIGHT = [
    # Preferido (según tu estructura): "Comp. - F. Emisión"
    ("comp_f_emision", 120),
    ("fecha_emision", 110),
    ("fecha_venta", 100),
    ("fecha_factura", 90),
    ("fecha_remito", 80),
    ("fecha", 50),
]


def _pick_serial_col(headers: list) -> Optional[int]:
    # Match exact preferred header first
    preferred = "item_articulo_partida_cod"
    for i, h in enumerate(headers):
        if h == preferred:
            return i
    # then contains for all hints
    for i, h in enumerate(headers):
        for key in SERIAL_HINTS:
            if key in h:
                return i
    return None


def _pick_date_cols(headers: list) -> list:
    cols = []
    for i, h in enumerate(headers):
        for key, weight in DATE_HINTS_WEIGHT:
            if key in h:
                cols.append((i, weight))
                break
    # sort by weight desc
    cols.sort(key=lambda x: x[1], reverse=True)
    return [i for i, _w in cols]


def find_serial_sale_date(numero_serie: str, brand_hint: Optional[str] = None) -> Tuple[Optional[_dt.date], Optional[Dict[str, Any]]]:
    """Busca el N/S en los Excels de trazabilidad y retorna la fecha de venta más reciente.

    Devuelve (fecha, meta) donde meta incluye {file, sheet, serial_value, date_column}.
    """
    if not numero_serie:
        return None, None
    target = _norm_serial(numero_serie)

    # Cache por N/S (normalizado)
    now = _dt.datetime.now().timestamp()
    cache_key = target + (f"|{_norm_txt(brand_hint)}" if brand_hint else "")
    cached = _CACHE.get(cache_key)
    if cached and (now - cached[2] <= _CACHE_TTL):
        return cached[0], cached[1]

    files = list_excel_files(brand_hint)
    if not files:
        return None, None

    last_date = None
    last_meta = None

    for path in files:
        ext = os.path.splitext(path)[1].lower()
        try:
            if ext == ".xlsx":
                it = _iter_rows_from_xlsx(path)
            elif ext == ".xls":
                it = _iter_rows_from_xls(path)
            else:
                continue

            for sheet_name, headers, rows in it:
                if not headers:
                    continue
                s_col = _pick_serial_col(headers)
                if s_col is None:
                    continue
                d_cols = _pick_date_cols(headers)
                if not d_cols:
                    continue
                for row in rows:
                    try:
                        raw_s = row[s_col] if s_col < len(row) else None
                    except Exception:
                        raw_s = None
                    if raw_s is None:
                        continue
                    s_norm = _norm_serial(raw_s)
                    if not (s_norm == target or (target and target in s_norm)):
                        continue
                    # matched serial; resolve best date from date columns
                    best = None
                    for dc in d_cols:
                        try:
                            raw_d = row[dc] if dc < len(row) else None
                        except Exception:
                            raw_d = None
                        dd = _parse_date(raw_d)
                        if dd and (best is None or dd > best):
                            best = dd
                    if best:
                        if last_date is None or best > last_date:
                            last_date = best
                            last_meta = {
                                "file": path,
                                "sheet": sheet_name,
                                "serial_value": str(raw_s),
                                "date": best.isoformat(),
                            }
        except Exception:
            # Si no se puede abrir/leer, continuar con el siguiente
            continue

    if last_date:
        _CACHE[cache_key] = (last_date, last_meta, now)
    return last_date, last_meta
_CACHE_TTL = int(os.getenv("TRAZABILIDAD_CACHE_TTL_SEC", "600"))  # 10 minutos por defecto
_CACHE: Dict[str, Tuple[_dt.date, Dict[str, Any], float]] = {}
