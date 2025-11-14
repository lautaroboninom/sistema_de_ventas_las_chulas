import os
import datetime as _dt
from typing import Optional, Dict, Any, Tuple

from django.conf import settings
from django.db import connection


def _norm_serial(s: str) -> str:
    try:
        import re
        s = str(s or "").strip().upper()
        s = re.sub(r"[\s\-_/]", "", s)
        return s
    except Exception:
        return (s or "").strip().upper()


def _parse_date(val) -> Optional[_dt.date]:
    if not val:
        return None
    if isinstance(val, _dt.datetime):
        return val.date()
    if isinstance(val, _dt.date):
        return val
    # Excel serial (heurística)
    if isinstance(val, (int, float)):
        try:
            if 20000 <= float(val) <= 60000:
                base = _dt.date(1899, 12, 30)  # convención Excel (Windows)
                return base + _dt.timedelta(days=float(val))
        except Exception:
            pass
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%m/%d/%Y", "%d.%m.%Y"):
        try:
            return _dt.datetime.strptime(s, fmt).date()
        except Exception:
            pass
    try:
        return _dt.date.fromisoformat(s)
    except Exception:
        return None


def _read_sale_date_from_general(numero_serie: str) -> Tuple[Optional[_dt.date], Optional[Dict[str, Any]]]:
    """Lee el archivo @GENERAL.xlsx y busca la fecha de venta en columna F
    según el número de serie exacto (normalizado) en columna E. Devuelve la
    fecha más reciente y metadatos básicos.
    """
    if not numero_serie:
        return None, None
    target = _norm_serial(numero_serie)

    xlsx_path = getattr(settings, "TRAZABILIDAD_GENERAL_FILE", None)
    if not xlsx_path or not os.path.isfile(xlsx_path):
        return None, None
    try:
        from openpyxl import load_workbook
    except Exception:
        return None, None

    last_date: Optional[_dt.date] = None
    last_meta: Optional[Dict[str, Any]] = None

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception:
        return None, None
    try:
        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
            except Exception:
                continue
            # Iteramos todas las filas; columnas E (idx 4) y F (idx 5)
            try:
                for row in ws.iter_rows(values_only=True):
                    try:
                        s_val = row[4] if len(row) > 4 else None
                        d_val = row[5] if len(row) > 5 else None
                    except Exception:
                        s_val, d_val = None, None
                    if s_val is None:
                        continue
                    if _norm_serial(s_val) != target:
                        continue
                    dd = _parse_date(d_val)
                    if dd and (last_date is None or dd > last_date):
                        last_date = dd
                        last_meta = {
                            "file": xlsx_path,
                            "sheet": sheet_name,
                            "serial_value": str(s_val),
                            "date": dd.isoformat(),
                        }
            except Exception:
                # Hoja problemática: continuar
                continue
    finally:
        try:
            wb.close()
        except Exception:
            pass

    return last_date, last_meta


def _rule_days_for(brand_id: Optional[int], model_id: Optional[int], serial_norm: str) -> Optional[int]:
    """Obtiene días de garantía desde warranty_rules si existe alguna regla activa.
    Prioridad: model_id > brand_id > serial_prefix. Devuelve None si no hay reglas.
    """
    try:
        with connection.cursor() as cur:
            # 1) Por modelo
            if model_id is not None:
                cur.execute(
                    """
                    SELECT days FROM warranty_rules
                     WHERE activo = TRUE AND model_id = %s
                     ORDER BY id DESC LIMIT 1
                    """,
                    [int(model_id)],
                )
                r = cur.fetchone()
                if r:
                    return int(r[0])
            # 2) Por marca
            if brand_id is not None:
                cur.execute(
                    """
                    SELECT days FROM warranty_rules
                     WHERE activo = TRUE AND brand_id = %s AND model_id IS NULL
                     ORDER BY id DESC LIMIT 1
                    """,
                    [int(brand_id)],
                )
                r = cur.fetchone()
                if r:
                    return int(r[0])
            # 3) Por prefijo de serie
            if serial_norm:
                cur.execute(
                    """
                    SELECT days FROM warranty_rules
                     WHERE activo = TRUE AND serial_prefix IS NOT NULL
                       AND %s LIKE (serial_prefix || '%')
                     ORDER BY length(serial_prefix) DESC, id DESC LIMIT 1
                    """,
                    [serial_norm],
                )
                r = cur.fetchone()
                if r:
                    return int(r[0])
    except Exception:
        return None
    return None


def compute_warranty(numero_serie: str,
                     brand_id: Optional[int] = None,
                     model_id: Optional[int] = None
                     ) -> Dict[str, Any]:
    """Cálculo de garantía usando reglas (si existen) y Excel GENERAL.

    Regla base: 365 días si no hay una regla. Si no hay fecha de venta
    en Excel, no se puede determinar; se devuelve garantia=None y vence=None.
    """
    today = _dt.date.today()
    serial_norm = _norm_serial(numero_serie)

    # Días por regla (si existe), sino default 365
    days = _rule_days_for(brand_id, model_id, serial_norm)
    if days is None:
        days = 365

    sale_date, meta = _read_sale_date_from_general(numero_serie)
    if not sale_date:
        return {
            "garantia": None,  # indeterminado por falta de fecha
            "vence_el": None,
            "fecha_venta": None,
            "days": days,
            "meta": meta or {"source": "excel_general", "file": getattr(settings, "TRAZABILIDAD_GENERAL_FILE", None)},
        }

    vence = sale_date + _dt.timedelta(days=int(days))
    en_garantia = today <= vence
    return {
        "garantia": bool(en_garantia),
        "vence_el": vence,
        "fecha_venta": sale_date,
        "days": days,
        "meta": meta or {"source": "excel_general", "file": getattr(settings, "TRAZABILIDAD_GENERAL_FILE", None)},
    }

